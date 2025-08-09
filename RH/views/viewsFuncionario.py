from django.shortcuts import render, HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.urls import reverse
from django.template.loader import render_to_string
from django.views import View
from ..models import Funcionario, Empresa, Lotacao, Cargo, ConfigGeral, Auditoria
from PerfilMenus.views import AcessoAcoes
from django.contrib.auth.models import User, Group
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..forms import FuncionarioForm
import json
from django.db import connection
# Create your views here.



class IndexView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return render(request, template_name = 'index.html')


class ListFuncionarioView(LoginRequiredMixin, View):
    template_name = 'Funcionario/listFuncionario.html'
    

    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Listar')
        if acesso == False: 
            return render(request, 'Forbidden.html')
        
        func = Funcionario.objects.filter(deleted='N')
        emp = Empresa.objects.values().distinct()
        lot = Lotacao.objects.values().distinct()
        car = Cargo.objects.values().distinct()
        
        # Aplicar filtros se fornecidos
        empresa_filter = request.GET.get('empresa')
        lotacao_filter = request.GET.get('lotacao')
        cargo_filter = request.GET.get('cargo')
        status_filter = request.GET.get('status')
        filtro_especial = request.GET.get('filtro_especial')
        
        if empresa_filter and empresa_filter != 'Todos':
            func = func.filter(empresa__empresa=empresa_filter)
        
        if lotacao_filter and lotacao_filter != 'Todos':
            func = func.filter(lotacao__lotacao=lotacao_filter)
        
        if cargo_filter and cargo_filter != 'Todos':
            func = func.filter(cargo__cargo=cargo_filter)
        
        if status_filter and status_filter != 'Todos':
            if status_filter == 'Ativo':
                func = func.filter(status='A')
            elif status_filter == 'Demitido':
                func = func.filter(status='D')
            elif status_filter == 'Afastado':
                func = func.filter(status='F')
        
        # Aplicar filtros especiais
        if filtro_especial and filtro_especial != 'Todos':
            from datetime import date, timedelta
            hoje = date.today()
            
            if filtro_especial == '1':  # Niver do mês
                func = func.filter(dt_nascimento__month=hoje.month)
            elif filtro_especial == '2':  # Venc. Exp 1 (7 dias)
                data_limite = hoje + timedelta(days=7)
                func = func.filter(
                    dt_primeiro_termino__gte=hoje,
                    dt_primeiro_termino__lte=data_limite
                )
            elif filtro_especial == '3':  # Venc. Exp 2 (7 dias)
                data_limite = hoje + timedelta(days=7)
                func = func.filter(
                    dt_segundo_termino__gte=hoje,
                    dt_segundo_termino__lte=data_limite
                )
            elif filtro_especial == '4':  # Venc. Exp 1 (30 dias)
                data_limite = hoje + timedelta(days=30)
                func = func.filter(
                    dt_primeiro_termino__gte=hoje,
                    dt_primeiro_termino__lte=data_limite
                )
            elif filtro_especial == '5':  # Venc. Exp 2 (30 dias)
                data_limite = hoje + timedelta(days=30)
                func = func.filter(
                    dt_segundo_termino__gte=hoje,
                    dt_segundo_termino__lte=data_limite
                )
        
        context = {
            'funcionario': func,
            'empresa': emp,
            'cargo': car,
            'lotacao': lot,
            'filtros': {
                'empresa': empresa_filter or '',
                'lotacao': lotacao_filter or '',
                'cargo': cargo_filter or '',
                'status': status_filter or '',
                'filtro_especial': filtro_especial or ''
            }
        }
        return render(request, self.template_name, context)
    

class AddFuncionarioView(LoginRequiredMixin, View):
    template_name = 'Funcionario/addFuncionario.html'
    
    def get(self, request, *args, **kwargs):
        
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Inserir')
        if acesso == False:
            return render(request, 'Forbidden.html')
        
        config = ConfigGeral.objects.all().values('empresa','dias_primeiro_termino_exp', 'dias_segundo_termino_exp' )
        print(config)
        form = FuncionarioForm()
        context = {
            'form': form,
            'configGeral': config
        }
        
        print(context)
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Inserir')
        if acesso == False:
            return render(request, 'Forbidden.html')
        
        form = FuncionarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listFuncionario')  # Redireciona para a lista de clientes após adicionar
        context = {
            'form': form,
        }
        return render(request, self.template_name, context)
    
    
class EditFuncionarioView(LoginRequiredMixin, View):
    template_name = 'Funcionario/addFuncionario.html'
    
    def get(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Atualizar')
        if acesso == False:
            return render(request, 'Forbidden.html')
        
        funcionario = get_object_or_404(Funcionario, id=id)
        
        funcionario.dt_admissao = funcionario.dt_admissao.strftime('%Y-%m-%d') if funcionario.dt_admissao else None
        funcionario.dt_demissao = funcionario.dt_demissao.strftime('%Y-%m-%d') if funcionario.dt_demissao else None
        funcionario.dt_nascimento = funcionario.dt_nascimento.strftime('%Y-%m-%d') if funcionario.dt_nascimento else None
        funcionario.dt_aso_periodico = funcionario.dt_aso_periodico.strftime('%Y-%m-%d') if funcionario.dt_aso_periodico else None
        funcionario.dt_contrato_experiencia = funcionario.dt_contrato_experiencia.strftime('%Y-%m-%d') if funcionario.dt_contrato_experiencia else None
        funcionario.dt_ultimo_aso = funcionario.dt_ultimo_aso.strftime('%Y-%m-%d') if funcionario.dt_ultimo_aso else None
        funcionario.dt_primeiro_termino = funcionario.dt_primeiro_termino.strftime('%Y-%m-%d') if funcionario.dt_primeiro_termino else None
        funcionario.dt_segundo_termino = funcionario.dt_segundo_termino.strftime('%Y-%m-%d') if funcionario.dt_segundo_termino else None
        funcionario.dt_integracao = funcionario.dt_integracao.strftime('%Y-%m-%d') if funcionario.dt_integracao else None
        
        form = FuncionarioForm(instance=funcionario)
        context = {
            'form': form,
            'funcionario': funcionario,
        }
        return render(request, self.template_name, context)

    def post(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Atualizar')
        if acesso == False:
            return render(request, 'Forbidden.html')
        
        funcionario = get_object_or_404(Funcionario , id=id)
        form = FuncionarioForm(request.POST, instance=funcionario)
        if form.is_valid():
            dados = form.save(commit=False)
            
            print(dados.status ,' - ', dados.dt_demissao)
            
            # if dados.dt_demissao:
            #     dados.status = 'D'  
            # elif dados.status != 'D':
            #     dados.dt_demissao = None
            #     print(dados.status ,' - ', dados.dt_demissao)
            # elif dados.status == 'D' and dados.dt_demissao == None:
            #     print('Estou aqui')
            #     dados.dt_demissao = dt.today().date()
            
            form.save()
            return redirect('listFuncionario')  # Redireciona para a lista de clientes após editar
        else:
            print('Erro no update!!!')
        context = {
            'form': form,
            'funcionario': funcionario,
        }
        return render(request, self.template_name, context)
    
    
    
class DelFuncionarioView(LoginRequiredMixin, View):
    
       def post(self, request, id):
           
        acesso = AcessoAcoes(request, 'Funcionarios', 'Deletar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'}) 
            
            
        func = get_object_or_404(Funcionario, id=id)
        try:
            func.deleted = 'S'
            func.save()
            # Retorna uma resposta JSON indicando sucesso
            return JsonResponse({'status': 'success', 'message': 'Funcionário deletado com sucesso!'})
        except Exception as e:
            # Retorna uma resposta JSON indicando falha
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)  
        
        

class RelatorioExcelView(LoginRequiredMixin, View):

    def post(self, request):
        
        try:
            # Verificar permissão
            acesso = AcessoAcoes(request, 'Funcionarios', 'Exportar')
            print(f"Resultado da verificação de acesso: {acesso}")
            
            if not acesso:
                print("Usuário sem permissão para exportar")
                return JsonResponse({
                    'status': 'forbidden', 
                    'message': 'Usuário sem permissão para exportar funcionários. Verifique se a ação "Exportar" está configurada para o menu "Funcionarios" e se o usuário tem permissão.'
                }) 
            
            emp = request.POST.get('emp', 'Todos')
            lot = request.POST.get('lot', 'Todos')
            car = request.POST.get('car', 'Todos')
            stt = request.POST.get('stt', 'Todos')
            
            print(f"Parâmetros recebidos: emp={emp}, lot={lot}, car={car}, stt={stt}")
            
            with connection.cursor() as cursor:
                # Query SQL parametrizada para evitar SQL injection
                query = """
                    SELECT 
                        f.id,
                        E.empresa,
                        f.matricula,
                        CONVERT(varchar(10), f.dt_admissao, 103) as dt_admissao,
                        CONVERT(varchar(10), f.dt_demissao, 103) as dt_demissao,
                        f.status,
                        f.nome AS Funcionário,
                        CONVERT(varchar(10), f.dt_nascimento, 103) as dt_nascimento,
                        l.lotacao,
                        tc.tipo_contrato,
                        c.cargo,
                        tu.turno,
                        f.sexo,
                        f.naturalidade,
                        f.rg,
                        f.cpf,
                        f.conta_bancaria,
                        f.motivo_contratacao,
                        CONVERT(varchar(10), f.dt_primeiro_termino, 103) as dt_primeiro_termino,
                        CONVERT(varchar(10), f.dt_segundo_termino, 103) as dt_segundo_termino,
                        CONVERT(varchar(10), f.dt_contrato_experiencia, 103) as dt_contrato_experiencia,
                        CONVERT(varchar(10), f.dt_ultimo_aso, 103) as dt_ultimo_aso,
                        CONVERT(varchar(10), f.dt_aso_periodico, 103) as dt_aso_periodico,
                        f.fone_fixo,
                        f.fone_celular,
                        fo.folga,
                        f.plano_saude_titular,
                        f.plano_saude_dependente,
                        f.plano_odonto_titular,
                        f.plano_odonto_dependente,
                        f.vale_transporte,
                        f.salario_familia,
                        f.dependentes,
                        f.salario_fixo,
                        f.salario_compl 
                    FROM RH_funcionario f
                    LEFT JOIN rh_empresa e ON e.id = f.empresa_id
                    LEFT JOIN RH_lotacao l ON l.id = f.lotacao_id
                    LEFT JOIN RH_cargo c ON c.id = f.cargo_id
                    LEFT JOIN RH_tipocontrato tc ON tc.id = f.tipo_contrato_id
                    LEFT JOIN RH_turno tu ON tu.id = f.turno_id
                    LEFT JOIN RH_folga fo ON fo.id = f.folga_id
                    WHERE f.deleted = 'N'
                """
                
                # Parâmetros para a query
                params = []
                
                # Adiciona filtros condicionalmente
                if emp != 'Todos':
                    query += " AND e.empresa = %s"
                    params.append(emp)
                
                if lot != 'Todos':
                    query += " AND l.lotacao = %s"
                    params.append(lot)
                
                if car != 'Todos':
                    query += " AND c.cargo = %s"
                    params.append(car)
                
                if stt != 'Todos':
                    if stt == 'Ativo':
                        query += " AND f.status = 'A'"
                    elif stt == 'Demitido':
                        query += " AND f.status = 'D'"
                    elif stt == 'Afastado':
                        query += " AND f.status = 'F'"
                
                print(f"Executando query: {query}")
                print(f"Parâmetros: {params}")
                
                cursor.execute(query, params)
                rows = cursor.fetchall()
                
                print(f"Registros encontrados: {len(rows)}")

            data = []
            for row in rows:
                data.append(row)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Lista de Funcionários"
            
            sheet.merge_cells("A1:D1")
            sheet["A1"] = "Lista de Funcionários"
            sheet["A1"].font = Font(size=16, name="Calibri", bold=True)
            
            fontFilter = Font(size=10, bold=True, name="Arial")
            sheet["A3"] = "Empresa:"
            sheet["A3"].font = fontFilter
            sheet["B3"] = emp
            sheet["A4"] = "Lotação:"
            sheet["A4"].font = fontFilter
            sheet["B4"] = lot
            sheet["A5"] = "Cargo:"
            sheet["A5"].font = fontFilter
            sheet["B5"] = car
            sheet["A6"] = "Status:"
            sheet["A6"].font = fontFilter
            sheet["B6"] = stt
            
            sheet["A7"] = ""
            
            # Cabeçalhos das colunas
            headers = ['Id','Empresa','Matricula','Dt_Admissao','Dt_Demissao','Status','Funcionário','Dt_Nascimento','Lotacao','Tipo_Contrato','Cargo','Turno','Sexo','Naturalidade','RG','CPF','Conta_Bancaria',
                    'Motivo_Contratacao','Dt_Primeiro_Termino','Dt_Segundo_Termino','Dt_Contrato_Experiencia','Dt_Ultimo_Aso','Dt_Aso_Periodico','Fone_Fixo','Fone_Celular','Folga','Plano_Saude_Titular',
                    'Plano_Saude_Dependente','Plano_Odonto_Titular','Plano_Odonto_Dependente','Vale_Transporte','Salario_Familia','Dependentes','Salario_Fixo','Salario_Compl']
        
            # Adiciona cabeçalhos
            sheet.append(headers)
            
            fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
            fontColumn = Font(size=10, name="Arial", color="F8F8FF", bold=True)
            for row in sheet["A8:AI8"]:
               for cell in row:
                   cell.fill = fill
                   cell.font = fontColumn
                   cell.alignment = Alignment(horizontal="center", vertical="center")
                               
            # Adiciona os dados
            for row in data:
                sheet.append(row)

            # Cria o arquivo para download
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="funcionarios.xlsx"'
            
            workbook.save(response)
            return response        
            
        except Exception as e:
            print(f"Erro ao gerar Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'status': 'error', 'message': f'Erro ao gerar o arquivo Excel: {str(e)}'}, status=500)
    
    
class ConfigGeralEmpresaView(LoginRequiredMixin, View):

    def post(self, request, id):

        try:
            config = get_object_or_404(ConfigGeral, Empresa=id)
            
            print(config)
           
            return HttpResponse(config)
        except Exception as e:
            # Retorna uma resposta JSON indicando falha
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)  
        
        

# def relatorioAuditoria(request):
#     data_atual = dt.date.today()
#     print(f"DATA ATUAL: {data_atual}")
#     template_name = 'core/controle/relatorio.html'
#     consulta_visitante = None
#     consulta_chave = None
#     consulta_veiculo = None
#     consulta_auditoria = None
#     consulta_funcionario = None
#     if request.method == 'POST':
#         opcao_filtro = request.POST.get('inlineRadioOptions')
#         data_inicial = request.POST.get('data_inicial')
#         data_final = request.POST.get('data_final')
#         print(opcao_filtro)
#         if opcao_filtro == "ControleVisitante":
#             consulta_visitante = ControlePessoa.objects.filter(
#                 pessoa__tipo__tipo='V', registro_entrada__gte=f"{data_inicial} 00:00:00", registro_entrada__lte=f"{data_final} 23:59:59")
#         elif opcao_filtro == "ControleFuncionario":
#             consulta_funcionario = ControlePessoa.objects.filter(
#                 pessoa__tipo__tipo='F', registro_entrada__gte=f"{data_inicial} 00:00:00", registro_entrada__lte=f"{data_final} 23:59:59")
#         elif opcao_filtro == "ControleChave":
#             consulta_chave = ControleChave.objects.filter(
#                 registro_entrega__gte=f"{data_inicial} 00:00:00", registro_entrega__lte=f"{data_final} 23:59:59")
#         elif opcao_filtro == "ControleVeiculo":
#             consulta_veiculo = ControleVeiculo.objects.filter(
#                 registro_entrada__gte=f"{data_inicial} 00:00:00", registro_entrada__lte=f"{data_final} 23:59:59")
#         elif opcao_filtro == "Auditoria":
#             consulta_auditoria = Auditoria.objects.filter(createDate__gte=f"{data_inicial} 00:00:00", createDate__lte=f"{data_final} 23:59:59")
        
#         context = {
#             'consulta_visitante': consulta_visitante,
#             'consulta_chave': consulta_chave,
#             'consulta_veiculo': consulta_veiculo,
#             'consulta_auditoria': consulta_auditoria,
#             'consulta_funcionario': consulta_funcionario,
#              'perfil': filtroPerfil(request),
#         }
#         print(consulta_funcionario)
#         return render(request, template_name, context)
#     else:
#         context = {
#             'dt_atual': data_atual,
#             'consulta_visitante': consulta_visitante,
#             'consulta_chave': consulta_chave,
#             'consulta_veiculo': consulta_veiculo,
#              'perfil': filtroPerfil(request),
#         }
#         return render(request, template_name, context)
            
def insertAuditoria(origem, user, obs):
    data = Auditoria(origem=origem, usuario=user, observacao=obs)
    data.save()            