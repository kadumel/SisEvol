from django.shortcuts import render, HttpResponse, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.urls import reverse
from django.template.loader import render_to_string
from django.views import View
from datetime import datetime, timedelta
from ..models import Funcionario, Empresa, Lotacao, Cargo, ConfigGeral, Auditoria, Evento, TipoEvento, DiaEvento, ControleEvento, FrequenciaEvento, Absenteismo
from ..forms import EventoForm, TipoEventoForm
from PerfilMenus.views import AcessoAcoes
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..forms import FuncionarioForm
from django.db import connection
import json

# Create your views here.


class GestaoEventosView(LoginRequiredMixin, View):
    template_name = 'Eventos/gestaoEventos.html'

    def get(self, request, *args, **kwargs):
        
        print(30*'*','incio da view Gestao Eventos','*'*30)
        acesso = AcessoAcoes(request, 'Eventos', 'Listar')
        if acesso == False:
             return render(request, 'Forbidden.html')
        
        eventos = Evento.objects.all().order_by('-created')
        tipos_evento = TipoEvento.objects.all().order_by('tipo_evento')
        funcionarios = Funcionario.objects.filter(status='A', deleted='N').order_by('nome')
        empresas = Empresa.objects.all().order_by('empresa')
        
        # Debug: verificar se há dados
        print(f"Eventos encontrados: {eventos.count()}")
        print(f"Tipos de evento encontrados: {tipos_evento.count()}")
        print(f"Funcionários ativos encontrados: {funcionarios.count()}")
        print(f"Empresas encontradas: {empresas.count()}")
        
        context = {
            'eventos': eventos,
            'tipos_evento': tipos_evento,
            'funcionarios': funcionarios,
            'empresas': empresas,
        }
        return render(request, self.template_name, context)


class GestaoTiposEventoView(LoginRequiredMixin, View):
    template_name = 'Eventos/gestaoTiposEvento.html'

    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'TipoEventos', 'Listar')
        if acesso == False:
            return render(request, 'Forbidden.html')
        
        tipos_evento = TipoEvento.objects.all().order_by('tipo_evento')
        
        context = {
            'tipos_evento': tipos_evento,
        }
        return render(request, self.template_name, context)


class AddEventoView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Inserir')
        if acesso == False:
             return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        print("=== DEBUG ADD EVENTO ===")
        print(f"POST data: {request.POST}")
        print(f"User: {request.user}")
        
        form = EventoForm(request.POST)
        print(f"Form is valid: {form.is_valid()}")
        
        if form.is_valid():
            try:
                evento = form.save(commit=False)
                evento.usuario = request.user
                evento.save()
                print(f"Evento salvo com sucesso: {evento.id}")
                return JsonResponse({'status': 'success', 'message': 'Evento adicionado com sucesso!'})
            except Exception as e:
                print(f"Erro ao salvar evento: {str(e)}")
                return JsonResponse({'status': 'error', 'message': f'Erro ao salvar: {str(e)}'})
        else:
            print(f"Form errors: {form.errors}")
            return JsonResponse({'status': 'error', 'message': f'Dados inválidos: {form.errors}'})


class UpdateEventoView(LoginRequiredMixin, View):
    
    def post(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Atualizar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        evento = get_object_or_404(Evento, id=id)
        
        form = EventoForm(request.POST, instance=evento)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.usuario = request.user
            evento.save()
            return JsonResponse({'status': 'success', 'message': 'Evento atualizado com sucesso!'})
        else:
            return JsonResponse({'status': 'error', 'message': f'Dados inválidos: {form.errors}'})


class GetEventoView(LoginRequiredMixin, View):
    
    def get(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            evento = get_object_or_404(Evento, id=id)
            evento_data = {
                'tipo': evento.tipo.id,
                'descricao': evento.descricao,
                'obs': evento.Obs or '',
                'vagas': evento.vagas or ''
            }
            return JsonResponse({'status': 'success', 'evento': evento_data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class DeleteEventoView(LoginRequiredMixin, View):
    
    def post(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Deletar')
        print(acesso)
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            evento = get_object_or_404(Evento, id=id)
            evento.delete()
            return JsonResponse({'status': 'success', 'message': 'Evento excluído com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class AddTipoEventoView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'TipoEventos', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        form = TipoEventoForm(request.POST)
        if form.is_valid():
            tipo_evento = form.save()
            return JsonResponse({'status': 'success', 'message': 'Tipo de evento adicionado com sucesso!'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Dados inválidos. Verifique os campos obrigatórios.'})


class UpdateTipoEventoView(LoginRequiredMixin, View):
    
    def post(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'TipoEventos', 'Atualizar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        tipo_evento = get_object_or_404(TipoEvento, id=id)
        form = TipoEventoForm(request.POST, instance=tipo_evento)
        if form.is_valid():
            form.save()
            return JsonResponse({'status': 'success', 'message': 'Tipo de evento atualizado com sucesso!'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Dados inválidos. Verifique os campos obrigatórios.'})


class GetTipoEventoView(LoginRequiredMixin, View):
    
    def get(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'TipoEventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            tipo_evento = get_object_or_404(TipoEvento, id=id)
            tipo_evento_data = {
                'tipo_evento': tipo_evento.tipo_evento
            }
            return JsonResponse({'status': 'success', 'tipo_evento': tipo_evento_data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class DeleteTipoEventoView(LoginRequiredMixin, View):
    
    def post(self, request, id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'TipoEventos', 'Deletar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            tipo_evento = get_object_or_404(TipoEvento, id=id)
            
            # Verificar se existem eventos usando este tipo
            eventos_count = Evento.objects.filter(tipo=tipo_evento).count()
            if eventos_count > 0:
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Não é possível excluir este tipo de evento. Existem {eventos_count} evento(s) associado(s).'
                })
            
            tipo_evento.delete()
            return JsonResponse({'status': 'success', 'message': 'Tipo de evento excluído com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class ExportEventosExcelView(LoginRequiredMixin, View):
    
    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Eventos"
            
            # Cabeçalhos
            headers = ['Tipo', 'Descrição', 'Vagas', 'Observação', 'Data Criação']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Dados
            eventos = Evento.objects.all().order_by('-created')
            for row, evento in enumerate(eventos, 2):
                ws.cell(row=row, column=1, value=evento.tipo.tipo_evento if evento.tipo else '')
                ws.cell(row=row, column=2, value=evento.descricao)
                ws.cell(row=row, column=3, value=evento.vagas)
                ws.cell(row=row, column=4, value=evento.Obs)
                ws.cell(row=row, column=5, value=evento.created.strftime('%d/%m/%Y %H:%M') if evento.created else '')
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Criar resposta
            from django.http import HttpResponse
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=eventos.xlsx'
            wb.save(response)
            return response
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


# Views para Inscrições de Funcionários
class GetInscricoesEventoView(LoginRequiredMixin, View):
    
    def get(self, request, evento_id, *args, **kwargs):
        
        print(f"=== DEBUG: GetInscricoesEventoView chamada com evento_id: {evento_id} ===")
        print(f"=== DEBUG: Usuário: {request.user.username} ===")
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        print(f"=== DEBUG: Resultado do acesso: {acesso} ===")
        
        if acesso == False:
            print(f"=== DEBUG: Acesso negado para usuário {request.user.username} ===")
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            print(f"=== DEBUG: Buscando inscrições para evento_id: {evento_id} ===")
            
            # Verificar se o evento existe
            from ..models import Evento
            evento = Evento.objects.filter(id=evento_id).first()
            if not evento:
                print(f"=== DEBUG: Evento com ID {evento_id} não encontrado ===")
                return JsonResponse({'status': 'error', 'message': 'Evento não encontrado'})
            
            print(f"=== DEBUG: Evento encontrado: {evento.descricao} ===")
            
            inscricoes = ControleEvento.objects.filter(evento_id=evento_id).select_related('funcionario', 'funcionario__lotacao', 'funcionario__empresa')
            print(f"=== DEBUG: Inscrições encontradas: {inscricoes.count()} ===")
            
            # Debug: listar todas as inscrições
            for insc in inscricoes:
                print(f"=== DEBUG: Inscrição ID {insc.id} - Funcionário: {insc.funcionario.nome if insc.funcionario else 'N/A'} ===")
            
            inscricoes_data = []
            print(inscricoes)
            for inscricao in inscricoes:
                try:
                    inscricao_data = {
                        'id': inscricao.id,
                        'funcionario_nome': inscricao.funcionario.nome if inscricao.funcionario else 'N/A',
                        'funcionario_cpf': inscricao.funcionario.cpf if inscricao.funcionario else 'N/A',
                        'funcionario_setor': inscricao.funcionario.lotacao.lotacao  if inscricao.funcionario and inscricao.funcionario.lotacao else None,
                        'funcionario_empresa': inscricao.funcionario.empresa.empresa if inscricao.funcionario and inscricao.funcionario.empresa else None,
                        'funcionario_empresa_id': inscricao.funcionario.empresa.id if inscricao.funcionario and inscricao.funcionario.empresa else None,
                        'data_inscricao': inscricao.created.strftime('%d/%m/%Y %H:%M') if inscricao.created else '',
                        'status': 'confirmado',  # Por padrão, inscrição é confirmada
                        'observacoes': '',  # ControleEvento não tem observações
                    }
                    inscricoes_data.append(inscricao_data)
                    print(f"=== DEBUG: Inscrição {inscricao.id} - {inscricao.funcionario.nome if inscricao.funcionario else 'N/A'} ===")
                except Exception as e:
                    print(f"=== DEBUG: Erro ao processar inscrição {inscricao.id}: {str(e)} ===")
                    continue
            
            print(f"=== DEBUG: Total de inscrições processadas: {len(inscricoes_data)} ===")
            print(f"=== DEBUG: Dados finais: {inscricoes_data} ===")
            
            return JsonResponse({'status': 'success', 'inscricoes': inscricoes_data})
        except Exception as e:
            print(f"=== DEBUG: Erro ao buscar inscrições: {str(e)} ===")
            import traceback
            print(f"=== DEBUG: Traceback: {traceback.format_exc()} ===")
            return JsonResponse({'status': 'error', 'message': str(e)})


class AddInscricaoView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            
            evento_id = request.POST.get('evento')
            funcionario_id = request.POST.get('funcionario')
            
            print(f"=== DEBUG: Tentando criar inscrição - evento_id: {evento_id}, funcionario_id: {funcionario_id} ===")
            
            # Verificar se já existe inscrição
            inscricao_existente = ControleEvento.objects.filter(
                evento_id=evento_id, 
                funcionario_id=funcionario_id
            ).first()
            
            if inscricao_existente:
                print(f"=== DEBUG: Inscrição já existe - ID: {inscricao_existente.id} ===")
                return JsonResponse({'status': 'error', 'message': 'Funcionário já está inscrito neste evento!'})
            
            # Criar nova inscrição
            inscricao = ControleEvento.objects.create(
                evento_id=evento_id,
                funcionario_id=funcionario_id,
                usuario=request.user
            )
            
            print(f"=== DEBUG: Inscrição criada com sucesso - ID: {inscricao.id} ===")
            return JsonResponse({'status': 'success', 'message': 'Inscrição realizada com sucesso!'})
        except Exception as e:
            print(f"=== DEBUG: Erro ao criar inscrição: {str(e)} ===")
            return JsonResponse({'status': 'error', 'message': str(e)})


class UpdateInscricaoView(LoginRequiredMixin, View):
    
    def post(self, request, inscricao_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Atualizar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            
            inscricao = get_object_or_404(ControleEvento, id=inscricao_id)
            
            # Atualizar apenas o usuário que fez a modificação
            inscricao.usuario = request.user
            inscricao.save()
            
            return JsonResponse({'status': 'success', 'message': 'Inscrição atualizada com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class GetInscricaoView(LoginRequiredMixin, View):
    
    def get(self, request, inscricao_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            
            inscricao = get_object_or_404(ControleEvento, id=inscricao_id)
            inscricao_data = {
                'funcionario': inscricao.funcionario.id,
                'status': 'confirmado',  # Por padrão
                'observacoes': '',  # ControleEvento não tem observações
            }
            
            return JsonResponse({'status': 'success', 'inscricao': inscricao_data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class DeleteInscricaoView(LoginRequiredMixin, View):
    
    def post(self, request, inscricao_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Deletar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            
            inscricao = get_object_or_404(ControleEvento, id=inscricao_id)
            inscricao.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Inscrição excluída com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class ExportInscricoesExcelView(LoginRequiredMixin, View):
    
    def get(self, request, evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import ControleEvento
            
            evento = get_object_or_404(Evento, id=evento_id)
            inscricoes = ControleEvento.objects.filter(evento_id=evento_id).select_related('funcionario', 'funcionario__lotacao', 'funcionario__empresa')
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Inscrições - {evento.descricao}"
            
            # Cabeçalhos
            headers = ['Funcionário', 'CPF', 'Empresa', 'Setor', 'Data Inscrição', 'Status', 'Observações']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Dados
            for row, inscricao in enumerate(inscricoes, 2):
                ws.cell(row=row, column=1, value=inscricao.funcionario.nome)
                ws.cell(row=row, column=2, value=inscricao.funcionario.cpf)
                ws.cell(row=row, column=3, value=inscricao.funcionario.empresa.empresa if inscricao.funcionario.empresa else '')
                ws.cell(row=row, column=4, value=inscricao.funcionario.lotacao.lotacao if inscricao.funcionario.lotacao else '')
                ws.cell(row=row, column=5, value=inscricao.created.strftime('%d/%m/%Y %H:%M') if inscricao.created else '')
                ws.cell(row=row, column=6, value='Confirmado')  # Por padrão
                ws.cell(row=row, column=7, value='')  # ControleEvento não tem observações
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Criar resposta
            from django.http import HttpResponse
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=inscricoes_evento_{evento_id}.xlsx'
            wb.save(response)
            return response
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


# Views para Dias do Evento
class GetDiasEventoView(LoginRequiredMixin, View):
    
    def get(self, request, evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import DiaEvento
            
            print(f"=== DEBUG: Buscando dias para evento_id: {evento_id} ===")
            
            # Verificar se o evento existe
            evento = Evento.objects.filter(id=evento_id).first()
            if not evento:
                print(f"=== DEBUG: Evento com ID {evento_id} não encontrado ===")
                return JsonResponse({'status': 'error', 'message': 'Evento não encontrado'})
            
            print(f"=== DEBUG: Evento encontrado: {evento.descricao} ===")
            
            dias_evento = DiaEvento.objects.filter(evento_id=evento_id).order_by('data', 'hora_inicio')
            print(f"=== DEBUG: Dias encontrados: {dias_evento.count()} ===")
            
            dias_data = []
            for dia in dias_evento:
                try:
                    dia_data = {
                        'id': dia.id,
                        'data': dia.data.strftime('%d/%m/%Y') if dia.data else '',
                        'hora_inicio': dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else '',
                        'hora_fim': dia.hora_fim.strftime('%H:%M') if dia.hora_fim else '',
                        'observacoes': dia.observacoes if dia.observacoes else '',
                        'data_iso': dia.data.isoformat() if dia.data else '',
                        'hora_inicio_iso': dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else '',
                        'hora_fim_iso': dia.hora_fim.strftime('%H:%M') if dia.hora_fim else '',
                    }
                    dias_data.append(dia_data)
                    print(f"=== DEBUG: Dia {dia.id} - {dia.data} ===")
                except Exception as e:
                    print(f"=== DEBUG: Erro ao processar dia {dia.id}: {str(e)} ===")
                    continue
            
            print(f"=== DEBUG: Total de dias processados: {len(dias_data)} ===")
            return JsonResponse({'status': 'success', 'dias': dias_data})
        except Exception as e:
            print(f"=== DEBUG: Erro ao buscar dias: {str(e)} ===")
            import traceback
            print(f"=== DEBUG: Traceback: {traceback.format_exc()} ===")
            return JsonResponse({'status': 'error', 'message': str(e)})


class AddDiaEventoView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import DiaEvento
            from datetime import datetime
            
            evento_id = request.POST.get('evento')
            data_str = request.POST.get('data')
            hora_inicio_str = request.POST.get('hora_inicio')
            hora_fim_str = request.POST.get('hora_fim')
            observacoes = request.POST.get('observacoes', '')
            
            print(f"=== DEBUG: Tentando criar dia do evento - evento_id: {evento_id}, data: {data_str} ===")
            
            # Validar dados obrigatórios
            if not evento_id or not data_str:
                return JsonResponse({'status': 'error', 'message': 'Evento e data são obrigatórios!'})
            
            # Converter data
            try:
                data = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Data inválida!'})
            
            # Converter horas se fornecidas
            hora_inicio = None
            hora_fim = None
            
            if hora_inicio_str:
                try:
                    hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Hora de início inválida!'})
            
            if hora_fim_str:
                try:
                    hora_fim = datetime.strptime(hora_fim_str, '%H:%M').time()
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Hora de fim inválida!'})
            
            # Verificar se já existe dia para esta data
            dia_existente = DiaEvento.objects.filter(
                evento_id=evento_id, 
                data=data
            ).first()
            
            if dia_existente:
                print(f"=== DEBUG: Dia já existe - ID: {dia_existente.id} ===")
                return JsonResponse({'status': 'error', 'message': 'Já existe um dia cadastrado para esta data!'})
            
            # Criar novo dia do evento
            dia_evento = DiaEvento.objects.create(
                evento_id=evento_id,
                data=data,
                hora_inicio=hora_inicio,
                hora_fim=hora_fim,
                observacoes=observacoes,
                usuario=request.user
            )
            
            print(f"=== DEBUG: Dia do evento criado com sucesso - ID: {dia_evento.id} ===")
            return JsonResponse({'status': 'success', 'message': 'Dia do evento adicionado com sucesso!'})
        except Exception as e:
            print(f"=== DEBUG: Erro ao criar dia do evento: {str(e)} ===")
            return JsonResponse({'status': 'error', 'message': str(e)})


class UpdateDiaEventoView(LoginRequiredMixin, View):
    
    def post(self, request, dia_evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Atualizar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import DiaEvento
            from datetime import datetime
            
            dia_evento = get_object_or_404(DiaEvento, id=dia_evento_id)
            
            data_str = request.POST.get('data')
            hora_inicio_str = request.POST.get('hora_inicio')
            hora_fim_str = request.POST.get('hora_fim')
            observacoes = request.POST.get('observacoes', '')
            
            # Converter data se fornecida
            if data_str:
                try:
                    data = datetime.strptime(data_str, '%Y-%m-%d').date()
                    dia_evento.data = data
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Data inválida!'})
            
            # Converter horas se fornecidas
            if hora_inicio_str:
                try:
                    hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                    dia_evento.hora_inicio = hora_inicio
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Hora de início inválida!'})
            else:
                dia_evento.hora_inicio = None
            
            if hora_fim_str:
                try:
                    hora_fim = datetime.strptime(hora_fim_str, '%H:%M').time()
                    dia_evento.hora_fim = hora_fim
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Hora de fim inválida!'})
            else:
                dia_evento.hora_fim = None
            
            dia_evento.observacoes = observacoes
            dia_evento.usuario = request.user
            dia_evento.save()
            
            return JsonResponse({'status': 'success', 'message': 'Dia do evento atualizado com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class GetDiaEventoView(LoginRequiredMixin, View):
    
    def get(self, request, dia_evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            from ..models import DiaEvento
            
            dia_evento = get_object_or_404(DiaEvento, id=dia_evento_id)
            dia_evento_data = {
                'data': dia_evento.data.isoformat() if dia_evento.data else '',
                'hora_inicio': dia_evento.hora_inicio.strftime('%H:%M') if dia_evento.hora_inicio else '',
                'hora_fim': dia_evento.hora_fim.strftime('%H:%M') if dia_evento.hora_fim else '',
                'observacoes': dia_evento.observacoes if dia_evento.observacoes else '',
            }
            
            return JsonResponse({'status': 'success', 'dia_evento': dia_evento_data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class DeleteDiaEventoView(LoginRequiredMixin, View):
    
    def post(self, request, dia_evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Deletar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            dia_evento = get_object_or_404(DiaEvento, id=dia_evento_id)
            evento_id = dia_evento.evento.id
            dia_evento.delete()
            return JsonResponse({'status': 'success', 'message': 'Dia do evento excluído com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


# Views para FrequenciaEvento
class GetFrequenciasEventoView(LoginRequiredMixin, View):
    """View para listar frequências de um evento específico"""
    
    def get(self, request, evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            evento = get_object_or_404(Evento, id=evento_id)
            
            # Buscar todas as inscrições do evento
            inscricoes = ControleEvento.objects.filter(evento=evento)
            
            # Buscar todos os dias do evento
            dias_evento = DiaEvento.objects.filter(evento=evento).order_by('data')
            
            # Buscar todas as frequências relacionadas
            frequencias = FrequenciaEvento.objects.filter(
                controle_evento__evento=evento
            ).select_related(
                'controle_evento__funcionario',
                'dia_evento'
            ).order_by('dia_evento__data', 'controle_evento__funcionario__nome')
            
            # Organizar dados para o template
            dados_frequencias = []
            
            for inscricao in inscricoes:
                funcionario = inscricao.funcionario
                frequencias_funcionario = []
                
                for dia in dias_evento:
                    # Buscar frequência específica para este funcionário e dia
                    frequencia = frequencias.filter(
                        controle_evento=inscricao,
                        dia_evento=dia
                    ).first()
                    
                    if frequencia:
                        frequencias_funcionario.append({
                            'id': frequencia.id,
                            'dia_evento_id': dia.id,
                            'data': dia.data.strftime('%d/%m/%Y'),
                            'hora_inicio': dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else '',
                            'hora_fim': dia.hora_fim.strftime('%H:%M') if dia.hora_fim else '',
                            'status': frequencia.status,
                            'status_display': frequencia.get_status_display(),
                            'hora_entrada': frequencia.hora_entrada.strftime('%H:%M') if frequencia.hora_entrada else '',
                            'hora_saida': frequencia.hora_saida.strftime('%H:%M') if frequencia.hora_saida else '',
                            'observacoes': frequencia.observacoes or ''
                        })
                    else:
                        # Criar registro vazio para dias sem frequência
                        frequencias_funcionario.append({
                            'id': None,
                            'dia_evento_id': dia.id,
                            'data': dia.data.strftime('%d/%m/%Y'),
                            'hora_inicio': dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else '',
                            'hora_fim': dia.hora_fim.strftime('%H:%M') if dia.hora_fim else '',
                            'status': '',
                            'status_display': '',
                            'hora_entrada': '',
                            'hora_saida': '',
                            'observacoes': ''
                        })
                
                dados_frequencias.append({
                    'inscricao_id': inscricao.id,
                    'funcionario_id': funcionario.id,
                    'funcionario_nome': funcionario.nome,
                    'funcionario_matricula': funcionario.matricula or '',
                    'funcionario_cargo': funcionario.cargo.cargo if funcionario.cargo else '',
                    'funcionario_lotacao': funcionario.lotacao.lotacao if funcionario.lotacao else '',
                    'frequencias': frequencias_funcionario
                })
            
            # Preparar dados dos dias para o cabeçalho da tabela
            dias_tabela = []
            for dia in dias_evento:
                dias_tabela.append({
                    'id': dia.id,
                    'data': dia.data.strftime('%d/%m/%Y'),
                    'hora_inicio': dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else '',
                    'hora_fim': dia.hora_fim.strftime('%H:%M') if dia.hora_fim else '',
                    'observacoes': dia.observacoes or ''
                })
            
            context = {
                'evento': {
                    'id': evento.id,
                    'descricao': evento.descricao,
                    'tipo': evento.tipo.tipo_evento
                },
                'dias_tabela': dias_tabela,
                'frequencias': dados_frequencias,
                'total_inscritos': inscricoes.count(),
                'total_dias': dias_evento.count()
            }
            
            html = render_to_string('Eventos/partials/tabela_frequencias.html', context)
            return JsonResponse({
                'status': 'success',
                'html': html,
                'dados': context
            })
            
        except Exception as e:
            print(f"Erro ao buscar frequências: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Erro ao buscar frequências: {str(e)}'})


class AddFrequenciaView(LoginRequiredMixin, View):
    """View para adicionar/atualizar frequência"""
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            inscricao_id = request.POST.get('inscricao_id')
            dia_evento_id = request.POST.get('dia_evento_id')
            status = request.POST.get('status', 'P')
            hora_entrada = request.POST.get('hora_entrada', '')
            hora_saida = request.POST.get('hora_saida', '')
            observacoes = request.POST.get('observacoes', '')
            
            # Validar dados obrigatórios
            if not inscricao_id or not dia_evento_id:
                return JsonResponse({'status': 'error', 'message': 'Dados obrigatórios não fornecidos'})
            
            # Buscar inscrição e dia do evento
            inscricao = get_object_or_404(ControleEvento, id=inscricao_id)
            dia_evento = get_object_or_404(DiaEvento, id=dia_evento_id)
            
            # Verificar se já existe frequência para esta combinação
            frequencia, created = FrequenciaEvento.objects.get_or_create(
                controle_evento=inscricao,
                dia_evento=dia_evento,
                defaults={
                    'status': status,
                    'hora_entrada': hora_entrada if hora_entrada else None,
                    'hora_saida': hora_saida if hora_saida else None,
                    'observacoes': observacoes,
                    'usuario_registro': request.user
                }
            )
            
            if not created:
                # Atualizar frequência existente
                frequencia.status = status
                frequencia.hora_entrada = hora_entrada if hora_entrada else None
                frequencia.hora_saida = hora_saida if hora_saida else None
                frequencia.observacoes = observacoes
                frequencia.usuario_registro = request.user
                frequencia.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Frequência registrada com sucesso!',
                'frequencia_id': frequencia.id
            })
            
        except Exception as e:
            print(f"Erro ao salvar frequência: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Erro ao salvar frequência: {str(e)}'})


class UpdateFrequenciaView(LoginRequiredMixin, View):
    """View para atualizar frequência existente"""
    
    def post(self, request, frequencia_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Atualizar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            frequencia = get_object_or_404(FrequenciaEvento, id=frequencia_id)
            
            status = request.POST.get('status', frequencia.status)
            hora_entrada = request.POST.get('hora_entrada', '')
            hora_saida = request.POST.get('hora_saida', '')
            observacoes = request.POST.get('observacoes', '')
            
            frequencia.status = status
            frequencia.hora_entrada = hora_entrada if hora_entrada else None
            frequencia.hora_saida = hora_saida if hora_saida else None
            frequencia.observacoes = observacoes
            frequencia.usuario_registro = request.user
            frequencia.save()
            
            return JsonResponse({'status': 'success', 'message': 'Frequência atualizada com sucesso!'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erro ao atualizar frequência: {str(e)}'})


class GetFrequenciaView(LoginRequiredMixin, View):
    """View para buscar dados de uma frequência específica"""
    
    def get(self, request, frequencia_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            frequencia = get_object_or_404(FrequenciaEvento, id=frequencia_id)
            
            frequencia_data = {
                'id': frequencia.id,
                'inscricao_id': frequencia.controle_evento.id,
                'dia_evento_id': frequencia.dia_evento.id,
                'status': frequencia.status,
                'hora_entrada': frequencia.hora_entrada.strftime('%H:%M') if frequencia.hora_entrada else '',
                'hora_saida': frequencia.hora_saida.strftime('%H:%M') if frequencia.hora_saida else '',
                'observacoes': frequencia.observacoes or '',
                'funcionario_nome': frequencia.controle_evento.funcionario.nome,
                'data_evento': frequencia.dia_evento.data.strftime('%d/%m/%Y')
            }
            
            return JsonResponse({'status': 'success', 'frequencia': frequencia_data})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class DeleteFrequenciaView(LoginRequiredMixin, View):
    """View para excluir frequência"""
    
    def post(self, request, frequencia_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Deletar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            frequencia = get_object_or_404(FrequenciaEvento, id=frequencia_id)
            evento_id = frequencia.controle_evento.evento.id
            frequencia.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Frequência excluída com sucesso!'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


class ExportFrequenciasExcelView(LoginRequiredMixin, View):
    """View para exportar frequências para Excel"""
    
    def get(self, request, evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Consultar')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            evento = get_object_or_404(Evento, id=evento_id)
            
            # Buscar dados das frequências
            inscricoes = ControleEvento.objects.filter(evento=evento)
            dias_evento = DiaEvento.objects.filter(evento=evento).order_by('data')
            frequencias = FrequenciaEvento.objects.filter(
                controle_evento__evento=evento
            ).select_related(
                'controle_evento__funcionario',
                'dia_evento'
            )
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Frequências - {evento.descricao}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Cabeçalho
            headers = ['Funcionário', 'Matrícula', 'Cargo', 'Lotação']
            for dia in dias_evento:
                headers.append(f"{dia.data.strftime('%d/%m/%Y')}\n{dia.hora_inicio.strftime('%H:%M') if dia.hora_inicio else ''}-{dia.hora_fim.strftime('%H:%M') if dia.hora_fim else ''}")
            
            # Aplicar estilos ao cabeçalho
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Dados
            row = 2
            for inscricao in inscricoes:
                funcionario = inscricao.funcionario
                
                # Dados básicos do funcionário
                ws.cell(row=row, column=1, value=funcionario.nome)
                ws.cell(row=row, column=2, value=funcionario.matricula or '')
                ws.cell(row=row, column=3, value=funcionario.cargo.cargo if funcionario.cargo else '')
                ws.cell(row=row, column=4, value=funcionario.lotacao.lotacao if funcionario.lotacao else '')
                
                # Frequências por dia
                col = 5
                for dia in dias_evento:
                    frequencia = frequencias.filter(
                        controle_evento=inscricao,
                        dia_evento=dia
                    ).first()
                    
                    if frequencia:
                        status_text = frequencia.get_status_display()
                        if frequencia.hora_entrada:
                            status_text += f" ({frequencia.hora_entrada.strftime('%H:%M')}"
                            if frequencia.hora_saida:
                                status_text += f"-{frequencia.hora_saida.strftime('%H:%M')}"
                            status_text += ")"
                        
                        ws.cell(row=row, column=col, value=status_text)
                    else:
                        ws.cell(row=row, column=col, value="Não registrado")
                    
                    col += 1
                
                row += 1
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Criar resposta
            from django.http import HttpResponse
            from datetime import datetime
            
            filename = f"frequencias_{evento.descricao.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erro ao exportar: {str(e)}'})


class SalvarFrequenciasLoteView(LoginRequiredMixin, View):
    """View para salvar frequências em lote"""
    
    def post(self, request, evento_id, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Eventos', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            evento = get_object_or_404(Evento, id=evento_id)
            frequencias_data = json.loads(request.POST.get('frequencias', '[]'))
            
            print(f"=== DEBUG: Salvando {len(frequencias_data)} frequências em lote")
            
            if not frequencias_data:
                return JsonResponse({'status': 'error', 'message': 'Nenhuma frequência fornecida'})
            
            # Contadores para feedback
            criadas = 0
            atualizadas = 0
            erros = 0
            
            for freq_data in frequencias_data:
                try:
                    inscricao_id = freq_data.get('inscricao_id')
                    dia_evento_id = freq_data.get('dia_evento_id')
                    status = freq_data.get('status', '')
                    
                    if not inscricao_id or not dia_evento_id or not status:
                        erros += 1
                        continue
                    
                    # Buscar inscrição e dia do evento
                    inscricao = get_object_or_404(ControleEvento, id=inscricao_id, evento=evento)
                    dia_evento = get_object_or_404(DiaEvento, id=dia_evento_id, evento=evento)
                    
                    # Verificar se já existe frequência para esta combinação
                    frequencia = FrequenciaEvento.objects.filter(
                        controle_evento=inscricao,
                        dia_evento=dia_evento
                    ).first()
                    
                    if frequencia:
                        # Atualizar frequência existente
                        frequencia.status = status
                        frequencia.usuario_registro = request.user
                        frequencia.save()
                        atualizadas += 1
                    else:
                        # Criar nova frequência apenas se status não estiver vazio
                        if status:
                            FrequenciaEvento.objects.create(
                                controle_evento=inscricao,
                                dia_evento=dia_evento,
                                status=status,
                                usuario_registro=request.user
                            )
                            criadas += 1
                        
                except Exception as e:
                    print(f"Erro ao processar frequência: {str(e)}")
                    erros += 1
            
            print(f"=== DEBUG: Resultado - Criadas: {criadas}, Atualizadas: {atualizadas}, Erros: {erros}")
            
            return JsonResponse({
                'status': 'success',
                'message': f'Frequências salvas com sucesso! Criadas: {criadas}, Atualizadas: {atualizadas}',
                'criadas': criadas,
                'atualizadas': atualizadas,
                'erros': erros
            })
            
        except Exception as e:
            print(f"Erro ao salvar frequências em lote: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Erro ao salvar frequências: {str(e)}'})


class GestaoAbsenteismoView(LoginRequiredMixin, View):
    """View para gerenciamento de absenteísmo"""
    template_name = 'RH/gestaoAbsenteismo.html'

    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Absenteismo', 'Listar')
        if acesso == False:
             return render(request, 'Forbidden.html')
        
        # Parâmetros de filtro
        empresa_filter = request.GET.getlist('empresa')
        empresa_desmarcadas = request.GET.get('empresa_desmarcadas')
        lotacao_id = request.GET.get('lotacao')
        cargo_id = request.GET.get('cargo')
        status_filter = request.GET.getlist('status')
        competencia = request.GET.get('competencia')
        
        # Definir competência padrão (mês atual)
        if not competencia:
            competencia = datetime.now().strftime('%Y-%m')
        
        # Definir status padrão (todos selecionados)
        if not status_filter:
            status_filter = ['A', 'F', 'D']
        
        # Buscar dados para os filtros
        empresas = Empresa.objects.all().order_by('empresa')
        lotacoes = Lotacao.objects.all().order_by('lotacao')
        cargos = Cargo.objects.all().order_by('cargo')
        
        # Filtrar funcionários
        funcionarios = Funcionario.objects.filter(deleted='N')
        if empresa_filter and not empresa_desmarcadas:
            funcionarios = funcionarios.filter(empresa_id__in=empresa_filter)
        if lotacao_id:
            funcionarios = funcionarios.filter(lotacao_id=lotacao_id)
        if cargo_id:
            funcionarios = funcionarios.filter(cargo_id=cargo_id)
        if status_filter:
            funcionarios = funcionarios.filter(status__in=status_filter)

        funcionarios = funcionarios.select_related('empresa', 'lotacao', 'cargo').order_by('nome')
        
        # Buscar registros de absenteísmo para a competência
        ano, mes = competencia.split('-')
        data_inicio = datetime(int(ano), int(mes), 1).date()
        if int(mes) == 12:
            data_fim = datetime(int(ano) + 1, 1, 1).date() - timedelta(days=1)
        else:
            data_fim = datetime(int(ano), int(mes) + 1, 1).date() - timedelta(days=1)
        
        # Buscar registros de absenteísmo
        registros_absenteismo = Absenteismo.objects.filter(
            data__gte=data_inicio,
            data__lte=data_fim
        ).select_related(
            'funcionario',
            'funcionario__empresa',
            'funcionario__lotacao',
            'funcionario__cargo'
        ).order_by('-data')
        
        # Aplicar filtros adicionais se necessário
        if empresa_filter:
            registros_absenteismo = registros_absenteismo.filter(funcionario__empresa_id__in=empresa_filter)
        if lotacao_id:
            registros_absenteismo = registros_absenteismo.filter(funcionario__lotacao_id=lotacao_id)
        if cargo_id:
            registros_absenteismo = registros_absenteismo.filter(funcionario__cargo_id=cargo_id)
        
        # Estatísticas baseadas no modelo Absenteismo
        total_funcionarios = funcionarios.count()
        total_registros = registros_absenteismo.count()
        
        # Calcular totais dos campos de absenteísmo
        total_dias_trabalho = sum(registro.dias_trabalho or 0 for registro in registros_absenteismo)
        total_dias_falta = sum(registro.dias_falta or 0 for registro in registros_absenteismo)
        total_dias_justificados = sum(registro.dias_justificados or 0 for registro in registros_absenteismo)
        total_dias_atraso = sum(registro.dias_atraso or 0 for registro in registros_absenteismo)
        total_dias_saida_antecipada = sum(registro.dias_saida_antecipada or 0 for registro in registros_absenteismo)
        total_dias_extras = sum(registro.dias_extras or 0 for registro in registros_absenteismo)
        total_dias_ferias = sum(registro.dias_ferias or 0 for registro in registros_absenteismo)
        
        # Calcular taxa de absenteísmo
        taxa_absenteismo = 0
        if total_dias_trabalho > 0:
            total_ausencias = total_dias_falta + total_dias_justificados
            taxa_absenteismo = (total_ausencias / total_dias_trabalho) * 100
        
        # Buscar funcionários com maior absenteísmo
        funcionarios_absenteismo = []
        for funcionario in funcionarios:
            registro = registros_absenteismo.filter(funcionario=funcionario).first()
            if registro:
                total_ausencias = (registro.dias_falta or 0) + (registro.dias_justificados or 0)
                dias_trabalho = registro.dias_trabalho or 0
                taxa_funcionario = (total_ausencias / dias_trabalho * 100) if dias_trabalho > 0 else 0
                
                funcionarios_absenteismo.append({
                    'funcionario': funcionario,
                    'registro': registro,
                    'total_ausencias': total_ausencias,
                    'taxa_absenteismo': taxa_funcionario
                })
        
        # Ordenar por taxa de absenteísmo (maior primeiro)
        funcionarios_absenteismo.sort(key=lambda x: x['taxa_absenteismo'], reverse=True)
        
        context = {
            'funcionarios': funcionarios,
            'empresas': empresas,
            'lotacoes': lotacoes,
            'cargos': cargos,
            'registros_absenteismo': registros_absenteismo,
            'funcionarios_absenteismo': funcionarios_absenteismo,
            'competencia': competencia,
            'estatisticas': {
                'total_funcionarios': total_funcionarios,
                'total_registros': total_registros,
                'total_dias_trabalho': total_dias_trabalho,
                'total_dias_falta': total_dias_falta,
                'total_dias_justificados': total_dias_justificados,
                'total_dias_atraso': total_dias_atraso,
                'total_dias_saida_antecipada': total_dias_saida_antecipada,
                'total_dias_extras': total_dias_extras,
                'total_dias_ferias': total_dias_ferias,
                'taxa_absenteismo': round(taxa_absenteismo, 2),
                'periodo_analise': f'Competência {competencia}'
            },
            'filtros': {
                'empresa': empresa_filter or [],
                'lotacao_id': lotacao_id,
                'cargo_id': cargo_id,
                'status': status_filter
            }
        }
        
        return render(request, self.template_name, context)


class CadastroAbsenteismoView(LoginRequiredMixin, View):
    """View para cadastro de absenteísmo em matriz por competência"""
    template_name = 'RH/cadastroAbsenteismo.html'

    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Absenteismo', 'Inserir')
        if acesso == False:
             return render(request, 'Forbidden.html')
        
        # Buscar dados para os filtros
        empresas = Empresa.objects.all().order_by('empresa')
        lotacoes = Lotacao.objects.all().order_by('lotacao')
        cargos = Cargo.objects.all().order_by('cargo')
        
        # Parâmetros de filtro
        empresa_filter = request.GET.getlist('empresa')
        lotacao_id = request.GET.get('lotacao')
        cargo_id = request.GET.get('cargo')
        competencia = request.GET.get('competencia')
        
        # Definir competência padrão (mês atual)
        if not competencia:
            competencia = datetime.now().strftime('%Y-%m')
        
        # Filtrar funcionários
        funcionarios = Funcionario.objects.filter(status='A', deleted='N')
        if empresa_filter:
            funcionarios = funcionarios.filter(empresa_id__in=empresa_filter)
        if lotacao_id:
            funcionarios = funcionarios.filter(lotacao_id=lotacao_id)
        if cargo_id:
            funcionarios = funcionarios.filter(cargo_id=cargo_id)
        
        funcionarios = funcionarios.select_related('empresa', 'lotacao', 'cargo').order_by('nome')
        
        # Buscar registros de absenteísmo existentes para a competência
        ano, mes = competencia.split('-')
        data_inicio = datetime(int(ano), int(mes), 1).date()
        if int(mes) == 12:
            data_fim = datetime(int(ano) + 1, 1, 1).date() - timedelta(days=1)
        else:
            data_fim = datetime(int(ano), int(mes) + 1, 1).date() - timedelta(days=1)
        
        # Buscar registros existentes para a competência
        registros_existentes = Absenteismo.objects.filter(
            data__gte=data_inicio,
            data__lte=data_fim
        ).select_related('funcionario')
        
        # Criar dicionário para facilitar busca (usando competência como chave)
        registros_dict = {}
        for registro in registros_existentes:
            # Usar o primeiro dia do mês como chave para consistência
            key = f"{str(registro.funcionario.id)}_{str(competencia)}"
            registros_dict[key] = registro
        
        context = {
            'empresas': empresas,
            'lotacoes': lotacoes,
            'cargos': cargos,
            'funcionarios': funcionarios,
            'competencia': competencia,
            'registros_dict': registros_dict,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'filtros': {
                'empresa': empresa_filter or [],
                'lotacao_id': lotacao_id,
                'cargo_id': cargo_id
            }
        }
        
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Absenteismo', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            funcionario_id = request.POST.get('funcionario_id')
            competencia = request.POST.get('competencia')
            dias_trabalho = request.POST.get('dias_trabalho', 0)
            dias_falta = request.POST.get('dias_falta', 0)
            dias_justificados = request.POST.get('dias_justificados', 0)
            dias_atraso = request.POST.get('dias_atraso', 0)
            dias_saida_antecipada = request.POST.get('dias_saida_antecipada', 0)
            dias_extras = request.POST.get('dias_extras', 0)
            dias_ferias = request.POST.get('dias_ferias', 0)
            
            # Validar dados obrigatórios
            if not funcionario_id or not competencia:
                return JsonResponse({'status': 'error', 'message': 'Dados obrigatórios não fornecidos'})
            
            # Buscar funcionário
            funcionario = get_object_or_404(Funcionario, id=funcionario_id)
            
            # Converter competência para data (primeiro dia do mês)
            ano, mes = competencia.split('-')
            data_registro = datetime(int(ano), int(mes), 1).date()
            
            # Verificar se já existe registro para este funcionário nesta competência
            registro_existente = Absenteismo.objects.filter(
                funcionario=funcionario,
                data=data_registro
            ).first()
            
            if registro_existente:
                # Atualizar registro existente
                registro_existente.dias_trabalho = dias_trabalho
                registro_existente.dias_falta = dias_falta
                registro_existente.dias_justificados = dias_justificados
                registro_existente.dias_atraso = dias_atraso
                registro_existente.dias_saida_antecipada = dias_saida_antecipada
                registro_existente.dias_extras = dias_extras
                registro_existente.dias_ferias = dias_ferias
                registro_existente.usuario = request.user
                registro_existente.save()
                message = 'Absenteísmo atualizado com sucesso!'
            else:
                # Criar novo registro
                Absenteismo.objects.create(
                    funcionario=funcionario,
                    data=data_registro,
                    dias_trabalho=dias_trabalho,
                    dias_falta=dias_falta,
                    dias_justificados=dias_justificados,
                    dias_atraso=dias_atraso,
                    dias_saida_antecipada=dias_saida_antecipada,
                    dias_extras=dias_extras,
                    dias_ferias=dias_ferias,
                    usuario=request.user
                )
                message = 'Absenteísmo registrado com sucesso!'
            
            return JsonResponse({'status': 'success', 'message': message})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erro ao registrar absenteísmo: {str(e)}'})


class SalvarAbsenteismoLoteView(LoginRequiredMixin, View):
    """View para salvar múltiplos registros de absenteísmo de uma vez"""
    
    def post(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Absenteismo', 'Inserir')
        if acesso == False:
            return JsonResponse({'status': 'forbidden', 'message': 'Usuário sem permissão para essa ação!!!'})
        
        try:
            registros = request.POST.getlist('registros[]')
            registros_salvos = 0
            registros_atualizados = 0
            
            for registro_data in registros:
                if registro_data:
                    data = json.loads(registro_data)
                    
                    funcionario_id = data.get('funcionario_id')
                    competencia = data.get('competencia')
                    dias_trabalho = data.get('dias_trabalho', 0)
                    dias_falta = data.get('dias_falta', 0)
                    dias_justificados = data.get('dias_justificados', 0)
                    dias_atraso = data.get('dias_atraso', 0)
                    dias_saida_antecipada = data.get('dias_saida_antecipada', 0)
                    dias_extras = data.get('dias_extras', 0)
                    dias_ferias = data.get('dias_ferias', 0)
                    
                    if funcionario_id and competencia:
                        funcionario = get_object_or_404(Funcionario, id=funcionario_id)
                        
                        # Converter competência para data (primeiro dia do mês)
                        ano, mes = competencia.split('-')
                        data_registro = datetime(int(ano), int(mes), 1).date()
                        
                        # Verificar se já existe registro para este funcionário nesta competência
                        registro_existente = Absenteismo.objects.filter(
                            funcionario=funcionario,
                            data=data_registro
                        ).first()
                        
                        if registro_existente:
                            # Atualizar
                            registro_existente.dias_trabalho = dias_trabalho
                            registro_existente.dias_falta = dias_falta
                            registro_existente.dias_justificados = dias_justificados
                            registro_existente.dias_atraso = dias_atraso
                            registro_existente.dias_saida_antecipada = dias_saida_antecipada
                            registro_existente.dias_extras = dias_extras
                            registro_existente.dias_ferias = dias_ferias
                            registro_existente.usuario = request.user
                            registro_existente.save()
                            registros_atualizados += 1
                        else:
                            # Criar novo
                            Absenteismo.objects.create(
                                funcionario=funcionario,
                                data=data_registro,
                                dias_trabalho=dias_trabalho,
                                dias_falta=dias_falta,
                                dias_justificados=dias_justificados,
                                dias_atraso=dias_atraso,
                                dias_saida_antecipada=dias_saida_antecipada,
                                dias_extras=dias_extras,
                                dias_ferias=dias_ferias,
                                usuario=request.user
                            )
                            registros_salvos += 1
            
            message = f'Processamento concluído! {registros_salvos} registros criados e {registros_atualizados} atualizados.'
            return JsonResponse({'status': 'success', 'message': message})
            
        except Exception as e:
            print(f"Erro ao salvar frequências em lote: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Erro ao salvar frequências: {str(e)}'})


class DashboardRHView(LoginRequiredMixin, View):
    """View para dashboard de indicadores de RH"""
    template_name = 'RH/dashboardRH.html'

    def get(self, request, *args, **kwargs):
        
        acesso = AcessoAcoes(request, 'Funcionarios', 'Listar')
        if acesso == False:
             return render(request, 'Forbidden.html')
        
        # Parâmetros de filtro
        empresa_filter = request.GET.getlist('empresa')
        empresa_desmarcadas = request.GET.get('empresa_desmarcadas')
        lotacao_id = request.GET.get('lotacao')
        cargo_id = request.GET.get('cargo')
        status_filter = request.GET.getlist('status')
        competencia = request.GET.get('competencia')
        
        # Definir competência padrão (mês atual)
        if not competencia:
            competencia = datetime.now().strftime('%Y-%m')
        
        # Definir status padrão (todos selecionados)
        if not status_filter:
            status_filter = ['A', 'F', 'D']
        
        # Buscar dados para os filtros
        empresas = Empresa.objects.all().order_by('empresa')
        lotacoes = Lotacao.objects.all().order_by('lotacao')
        cargos = Cargo.objects.all().order_by('cargo')
        
        # Filtrar funcionários
        funcionarios = Funcionario.objects.filter(deleted='N')
        if empresa_filter and not empresa_desmarcadas:
            funcionarios = funcionarios.filter(empresa_id__in=empresa_filter)
        if lotacao_id:
            funcionarios = funcionarios.filter(lotacao_id=lotacao_id)
        if cargo_id:
            funcionarios = funcionarios.filter(cargo_id=cargo_id)
        if status_filter:
            funcionarios = funcionarios.filter(status__in=status_filter)

        funcionarios = funcionarios.select_related('empresa', 'lotacao', 'cargo').order_by('nome')
        
        # Calcular indicadores básicos
        total_funcionarios = funcionarios.count()
        funcionarios_ativos = funcionarios.filter(status='A').count()
        funcionarios_afastados = funcionarios.filter(status='F').count()
        funcionarios_demitidos = funcionarios.filter(status='D').count()
        
        # Indicadores por gênero
        funcionarios_masculino = funcionarios.filter(sexo='M').count()
        funcionarios_feminino = funcionarios.filter(sexo='F').count()
        funcionarios_outros = funcionarios.filter(sexo='T').count()
        
        # Calcular faixas etárias
        from datetime import date
        hoje = date.today()
        
        faixas_etarias = {
            '18-25': 0,
            '26-35': 0,
            '36-45': 0,
            '46-55': 0,
            '56-65': 0,
            '65+': 0
        }
        
        for funcionario in funcionarios:
            if funcionario.dt_nascimento:
                idade = hoje.year - funcionario.dt_nascimento.year
                if hoje.month < funcionario.dt_nascimento.month or (hoje.month == funcionario.dt_nascimento.month and hoje.day < funcionario.dt_nascimento.day):
                    idade -= 1
                
                if idade < 18:
                    continue
                elif idade <= 25:
                    faixas_etarias['18-25'] += 1
                elif idade <= 35:
                    faixas_etarias['26-35'] += 1
                elif idade <= 45:
                    faixas_etarias['36-45'] += 1
                elif idade <= 55:
                    faixas_etarias['46-55'] += 1
                elif idade <= 65:
                    faixas_etarias['56-65'] += 1
                else:
                    faixas_etarias['65+'] += 1
        
        # Calcular tempo médio de empresa
        tempo_medio_empresa = 0
        funcionarios_com_admissao = 0
        
        for funcionario in funcionarios.filter(status='A'):
            if funcionario.dt_admissao:
                tempo_empresa = hoje.year - funcionario.dt_admissao.year
                if hoje.month < funcionario.dt_admissao.month or (hoje.month == funcionario.dt_admissao.month and hoje.day < funcionario.dt_admissao.day):
                    tempo_empresa -= 1
                tempo_medio_empresa += tempo_empresa
                funcionarios_com_admissao += 1
        
        if funcionarios_com_admissao > 0:
            tempo_medio_empresa = round(tempo_medio_empresa / funcionarios_com_admissao, 1)
        
        # Calcular turnover (apenas se competência for relevante)
        turnover_mensal = 0
        turnover_anual = 0
        turnover_evolucao = []
        
        if competencia:
            ano, mes = competencia.split('-')
            data_inicio = datetime(int(ano), int(mes), 1).date()
            if int(mes) == 12:
                data_fim = datetime(int(ano) + 1, 1, 1).date() - timedelta(days=1)
            else:
                data_fim = datetime(int(ano), int(mes) + 1, 1).date() - timedelta(days=1)
            
            # Funcionários que entraram no mês
            admissoes_mes = funcionarios.filter(
                dt_admissao__gte=data_inicio,
                dt_admissao__lte=data_fim
            ).count()
            
            # Funcionários que saíram no mês
            demissoes_mes = funcionarios.filter(
                dt_demissao__gte=data_inicio,
                dt_demissao__lte=data_fim
            ).count()
            
            # Calcular turnover mensal
            if funcionarios_ativos > 0:
                turnover_mensal = round((demissoes_mes / funcionarios_ativos) * 100, 2)
            
            # Calcular turnover anual (últimos 12 meses)
            data_inicio_ano = datetime(int(ano), 1, 1).date()
            admissoes_ano = funcionarios.filter(
                dt_admissao__gte=data_inicio_ano,
                dt_admissao__lte=data_fim
            ).count()
            
            demissoes_ano = funcionarios.filter(
                dt_demissao__gte=data_inicio_ano,
                dt_demissao__lte=data_fim
            ).count()
            
            if funcionarios_ativos > 0:
                turnover_anual = round(((admissoes_ano + demissoes_ano) / 2 / funcionarios_ativos) * 100, 2)
            
            # Calcular evolução do turnover mês a mês do ano selecionado
            # Determinar até qual mês calcular baseado no ano atual
            ano_atual = datetime.now().year
            mes_atual = datetime.now().month
            
            # Se for o ano atual, calcular apenas até o mês atual
            # Se for ano anterior, calcular todos os 12 meses
            if int(ano) == ano_atual:
                meses_para_calcular = mes_atual
            else:
                meses_para_calcular = 12
            
            for mes_atual in range(1, meses_para_calcular + 1):
                # Data de início do mês
                data_inicio_mes = datetime(int(ano), mes_atual, 1).date()
                
                # Data de fim do mês
                if mes_atual == 12:
                    data_fim_mes = datetime(int(ano) + 1, 1, 1).date() - timedelta(days=1)
                else:
                    data_fim_mes = datetime(int(ano), mes_atual + 1, 1).date() - timedelta(days=1)
                
                # Calcular funcionários ativos no início do mês (aproximação)
                # Para simplificar, usamos o total atual de funcionários ativos
                # Em uma implementação mais precisa, seria necessário calcular o efetivo de cada mês
                funcionarios_ativos_mes = funcionarios_ativos
                
                # Funcionários que saíram no mês
                demissoes_mes_atual = funcionarios.filter(
                    dt_demissao__gte=data_inicio_mes,
                    dt_demissao__lte=data_fim_mes
                ).count()
                
                # Calcular turnover do mês
                turnover_mes = 0
                if funcionarios_ativos_mes > 0:
                    turnover_mes = round((demissoes_mes_atual / funcionarios_ativos_mes) * 100, 2)
                
                turnover_evolucao.append({
                    'mes': mes_atual,
                    'nome_mes': datetime(int(ano), mes_atual, 1).strftime('%b'),
                    'turnover': turnover_mes,
                    'demissoes': demissoes_mes_atual
                })
        
        # Indicadores por empresa
        funcionarios_por_empresa = {}
        for empresa in empresas:
            count = funcionarios.filter(empresa=empresa).count()
            if count > 0:
                funcionarios_por_empresa[empresa.empresa] = count
        
        # Indicadores por cargo
        funcionarios_por_cargo = {}
        for cargo in cargos:
            count = funcionarios.filter(cargo=cargo).count()
            if count > 0:
                funcionarios_por_cargo[cargo.cargo] = count
        
        # Indicadores por lotação
        funcionarios_por_lotacao = {}
        for lotacao in lotacoes:
            count = funcionarios.filter(lotacao=lotacao).count()
            if count > 0:
                funcionarios_por_lotacao[lotacao.lotacao] = count
        
        # Calcular absenteísmo médio (se competência for relevante)
        absenteismo_medio = 0
        if competencia:
            ano, mes = competencia.split('-')
            data_inicio = datetime(int(ano), int(mes), 1).date()
            if int(mes) == 12:
                data_fim = datetime(int(ano) + 1, 1, 1).date() - timedelta(days=1)
            else:
                data_fim = datetime(int(ano), int(mes) + 1, 1).date() - timedelta(days=1)
            
            registros_absenteismo = Absenteismo.objects.filter(
                data__gte=data_inicio,
                data__lte=data_fim,
                funcionario__in=funcionarios
            )
            
            total_dias_trabalho = sum(registro.dias_trabalho or 0 for registro in registros_absenteismo)
            total_dias_falta = sum(registro.dias_falta or 0 for registro in registros_absenteismo)
            total_dias_justificados = sum(registro.dias_justificados or 0 for registro in registros_absenteismo)
            
            if total_dias_trabalho > 0:
                absenteismo_medio = round(((total_dias_falta + total_dias_justificados) / total_dias_trabalho) * 100, 2)
        
        # Calcular indicadores de período de experiência
        funcionarios_em_experiencia = 0
        funcionarios_passaram_primeiro_termino = 0
        funcionarios_passaram_segundo_termino = 0
        funcionarios_contrato_experiencia = 0
        
        for funcionario in funcionarios.filter(status='A'):
            if funcionario.dt_admissao:
                # Verificar se passou do primeiro período
                if funcionario.dt_primeiro_termino and hoje > funcionario.dt_primeiro_termino:
                    funcionarios_passaram_primeiro_termino += 1
                
                # Verificar se passou do segundo período
                if funcionario.dt_segundo_termino and hoje > funcionario.dt_segundo_termino:
                    funcionarios_passaram_segundo_termino += 1
                
                # Verificar se ainda está no período de experiência
                if funcionario.dt_contrato_experiencia and hoje <= funcionario.dt_contrato_experiencia:
                    funcionarios_em_experiencia += 1
                
                # Verificar se tem contrato de experiência
                if funcionario.dt_contrato_experiencia:
                    funcionarios_contrato_experiencia += 1
        
        # Calcular percentuais
        percentual_passaram_primeiro = 0
        percentual_passaram_segundo = 0
        percentual_em_experiencia = 0
        
        if funcionarios_contrato_experiencia > 0:
            percentual_passaram_primeiro = round((funcionarios_passaram_primeiro_termino / funcionarios_contrato_experiencia) * 100, 1)
            percentual_passaram_segundo = round((funcionarios_passaram_segundo_termino / funcionarios_contrato_experiencia) * 100, 1)
            percentual_em_experiencia = round((funcionarios_em_experiencia / funcionarios_contrato_experiencia) * 100, 1)
        
        context = {
            'empresas': empresas,
            'lotacoes': lotacoes,
            'cargos': cargos,
            'competencia': competencia,
            'filtros': {
                'empresa': empresa_filter or [],
                'lotacao_id': lotacao_id or '',
                'cargo_id': cargo_id or '',
                'status': status_filter or [],
            },
            'indicadores': {
                'total_funcionarios': total_funcionarios,
                'funcionarios_ativos': funcionarios_ativos,
                'funcionarios_afastados': funcionarios_afastados,
                'funcionarios_demitidos': funcionarios_demitidos,
                'funcionarios_masculino': funcionarios_masculino,
                'funcionarios_feminino': funcionarios_feminino,
                'funcionarios_outros': funcionarios_outros,
                'faixas_etarias': faixas_etarias,
                'faixa_18_25': faixas_etarias['18-25'],
                'faixa_26_35': faixas_etarias['26-35'],
                'faixa_36_45': faixas_etarias['36-45'],
                'faixa_46_55': faixas_etarias['46-55'],
                'faixa_56_65': faixas_etarias['56-65'],
                'faixa_65_mais': faixas_etarias['65+'],
                'tempo_medio_empresa': tempo_medio_empresa,
                'turnover_mensal': turnover_mensal,
                'turnover_anual': turnover_anual,
                'turnover_evolucao': turnover_evolucao,
                'absenteismo_medio': absenteismo_medio,
                'funcionarios_por_empresa': funcionarios_por_empresa,
                'funcionarios_por_cargo': funcionarios_por_cargo,
                'funcionarios_por_lotacao': funcionarios_por_lotacao,
                'funcionarios_em_experiencia': funcionarios_em_experiencia,
                'funcionarios_passaram_primeiro_termino': funcionarios_passaram_primeiro_termino,
                'funcionarios_passaram_segundo_termino': funcionarios_passaram_segundo_termino,
                'funcionarios_contrato_experiencia': funcionarios_contrato_experiencia,
                'percentual_passaram_primeiro': percentual_passaram_primeiro,
                'percentual_passaram_segundo': percentual_passaram_segundo,
                'percentual_em_experiencia': percentual_em_experiencia,
            }
        }
        
        return render(request, self.template_name, context)

