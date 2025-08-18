from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import View
from django.db.models import Q, Sum
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
from .models import Orcamento, Conta, CentroCusto, Dre
from RH.models import Empresa
from django.db import connection
from PerfilMenus.views import AcessoAcoes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create your views here.

@require_POST
def update_orcamento_valor(request):
    """Atualiza valores de orçamento via AJAX"""
    # Verificar permissão para editar orçamento
    acesso = AcessoAcoes(request, 'Gestao Orçamento', 'Orçamento')
    if not acesso:
        return JsonResponse({
            'success': False,
            'error': 'Usuário sem permissão para editar orçamentos'
        })
    
    try:
        data = json.loads(request.body)
        
        # Verificar se é um valor único ou múltiplos valores
        if 'valores' in data:
            # Múltiplos valores (salvamento em lote)
            valores = data['valores']
            sucessos = 0
            erros = []
            
            for item in valores:
                orcamento_id = item.get('orcamento_id')
                valor = item.get('valor')
                
                if not orcamento_id or orcamento_id == 'None':
                    erros.append(f"ID de orçamento inválido: {orcamento_id}")
                    continue
                
                try:
                    orcamento = Orcamento.objects.get(id=orcamento_id)
                    orcamento.valor = valor
                    orcamento.save()
                    sucessos += 1
                except Orcamento.DoesNotExist:
                    erros.append(f"Orçamento não encontrado: {orcamento_id}")
                except Exception as e:
                    erros.append(f"Erro ao salvar orçamento {orcamento_id}: {str(e)}")
            
            if erros:
                return JsonResponse({
                    'success': False,
                    'error': f"Salvos: {sucessos}, Erros: {len(erros)}. Primeiro erro: {erros[0]}"
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': f"{sucessos} valores salvos com sucesso"
                })
        else:
            # Valor único (compatibilidade com versão anterior)
            orcamento_id = data.get('orcamento_id')
            valor = data.get('valor')
            
            if not orcamento_id or orcamento_id == 'None':
                return JsonResponse({
                    'success': False,
                    'error': 'ID de orçamento não fornecido'
                })
            
            try:
                orcamento = Orcamento.objects.get(id=orcamento_id)
                orcamento.valor = valor
                orcamento.save()
                return JsonResponse({'success': True})
            except Orcamento.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Orçamento não encontrado'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                })
                
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Dados JSON inválidos'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


class OrcamentoMenuView(LoginRequiredMixin, View):
    """View para o menu principal do orçamento"""
    template_name = 'Gestao/orcamento_menu.html'
    
    def get(self, request, *args, **kwargs):
        # Verificar permissão para acessar o menu de orçamento
        acesso = AcessoAcoes(request, 'Gestao Orçamento', 'Orçamento')
        if not acesso:
            return render(request, 'Forbidden.html')
        
        # Calcular estatísticas
        from .models import Conta, Orcamento
        from django.db.models import Sum
        
        total_empresas = Empresa.objects.count()
        total_contas = Conta.objects.count()
        total_orcamentos = Orcamento.objects.count()
        valor_total = Orcamento.objects.aggregate(total=Sum('valor'))['total'] or 0
        
        context = {
            'title': 'Menu Orçamento',
            'total_empresas': total_empresas,
            'total_contas': total_contas,
            'total_orcamentos': total_orcamentos,
            'valor_total': f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            'menu_items': [
                {
                    'title': 'DRE Orçado',
                    'description': 'Visualize o DRE Orçado com filtros hierárquicos',
                    'icon': 'bi bi-graph-up',
                    'url': 'dre_orcado',
                    'color': 'primary'
                },
                {
                    'title': 'Visualização Geral',
                    'description': 'Visualize o orçamento por empresa, ano e conta',
                    'icon': 'bi bi-table',
                    'url': 'orcamento_visualizacao',
                    'color': 'secondary'
                },
                {
                    'title': 'Administração',
                    'description': 'Gerencie orçamentos no painel administrativo',
                    'icon': 'bi bi-gear',
                    'url': 'admin:Gestao_orcamento_changelist',
                    'color': 'info'
                },
            ]
        }
        return render(request, self.template_name, context)


class DreOrcadoView(LoginRequiredMixin, View):
    """View para DRE Orçado com filtros e tabela hierárquica"""
    template_name = 'Gestao/dre_orcado.html'
    
    def get(self, request, *args, **kwargs):
        # Verificar permissão para acessar DRE Orçado
        acesso = AcessoAcoes(request, 'Gestao Orçamento', 'Orçamento')
        if not acesso:
            return render(request, 'Forbidden.html')
        
        # Parâmetros de filtro - suportando múltipla escolha
        ano = request.GET.get('ano', datetime.now().year)
        mes_list = request.GET.getlist('mes')  # Múltipla escolha para meses
        empresa_list = request.GET.getlist('empresa')  # Múltipla escolha para empresas
        filtro_livre = request.GET.get('filtro_livre', '').strip()
        
        # Buscar dados para os filtros
        # Filtrar apenas empresas que tenham movimentações no orçamento
        empresas = Empresa.objects.filter(
            id__in=Orcamento.objects.values_list('empresa_id', flat=True).distinct()
        ).order_by('empresa')
        anos = range(2020, datetime.now().year + 2)
        meses = [
            {'valor': '1', 'nome': 'Janeiro'}, {'valor': '2', 'nome': 'Fevereiro'},
            {'valor': '3', 'nome': 'Março'}, {'valor': '4', 'nome': 'Abril'},
            {'valor': '5', 'nome': 'Maio'}, {'valor': '6', 'nome': 'Junho'},
            {'valor': '7', 'nome': 'Julho'}, {'valor': '8', 'nome': 'Agosto'},
            {'valor': '9', 'nome': 'Setembro'}, {'valor': '10', 'nome': 'Outubro'},
            {'valor': '11', 'nome': 'Novembro'}, {'valor': '12', 'nome': 'Dezembro'}
        ]
        
        # Buscar estrutura DRE completa
        dres = Dre.objects.filter(id__lte=11).order_by('nivel')
        
        # Construir a consulta SQL
        sql = """
        select 
            eomonth(o.data) data,
            o.id idOrcamento,
            e.id cdEmpresa, e.empresa nmEmpresa, 
            d.id dre_id, d.nivel dre,
            case
                when c.nivel = 1  then c.codigo +' - '+ c.nome
                when c.nivel = 2 then (select codigo +' - '+ nome from SISEVOL..Gestao_conta n1 where n1.id = c.mae_id) 
                when c.nivel = 3 and nivel_dre_id is not null then (select n2.codigo +' - '+ n2.nome from SISEVOL..Gestao_conta n2 where n2.id = (select n2.mae_id from SISEVOL..Gestao_conta n2 where n2.id = c.nivel_dre_id ))
                when c.nivel = 3 then (select codigo +' - '+ nome from sisevol..Gestao_conta n1 where n1.id = (select n2.mae_id from SISEVOL..Gestao_conta n2 where n2.id = c.mae_id)) 
            end as nivel1,
            case 
                when c.nivel in ( 1, 2 ) then c.codigo +' - '+ c.nome
                when c.nivel = 3 and nivel_dre_id is not null then (select n2.codigo +' - '+ n2.nome from SISEVOL..Gestao_conta n2 where n2.id = c.nivel_dre_id )
                when c.nivel = 3 then (select codigo +' - '+ nome from SISEVOL..Gestao_conta n1 where n1.id = c.mae_id) 
            end as nivel2,
            c.codigo +' - '+ c.nome nivel3,
            convert(varchar(5), e.codigo_bi) + ' - ' + cr.codigo + ' - ' +  cr.nome CentroResultado,
            o.valor valor
        from Gestao_orcamento o
        join RH_empresa e on e.id = o.empresa_id
        join gestao_conta c on c.id = o.conta_id
        join gestao_dre d on d.id = c.dre_id
        join gestao_centrocusto cr on cr.id = o.centro_custo_id
        where 1=1
        """
        
        params = []
        
        # Aplicar filtros
        if ano:
            sql += " and year(o.data) = %s"
            params.append(ano)
        
        # Filtro de meses - múltipla escolha
        if mes_list:
            placeholders = ','.join(['%s'] * len(mes_list))
            sql += f" and month(o.data) in ({placeholders})"
            params.extend(mes_list)
        
        # Filtro de empresas - múltipla escolha
        if empresa_list:
            placeholders = ','.join(['%s'] * len(empresa_list))
            sql += f" and e.id in ({placeholders})"
            params.extend(empresa_list)
        
        if filtro_livre:
            sql += """ and (
                e.empresa like %s or 
                c.codigo like %s or 
                c.nome like %s or
                d.nivel like %s
            )"""
            filtro_param = f"%{filtro_livre}%"
            params.extend([filtro_param, filtro_param, filtro_param, filtro_param])
        
        sql += """
        order by d.nivel, nivel1, nivel2, nivel3, eomonth(o.data)
        """
        
        # Executar consulta
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            resultados = cursor.fetchall()
        
        # Processar resultados para estrutura hierárquica com datas
        dados_hierarquicos, datas_unicas = self.processar_dados_hierarquicos_com_datas(resultados)
        
        # Garantir que todas as DREs apareçam, mesmo sem dados
        dados_hierarquicos_completos = self.estrutura_completa_dre(dres, dados_hierarquicos, datas_unicas)
        
        context = {
            'empresas': empresas,
            'anos': anos,
            'meses': meses,
            'dres': dres,
            'dados_hierarquicos': dados_hierarquicos_completos,
            'datas_unicas': datas_unicas,
            'filtros': {
                'ano': ano,
                'mes': mes_list,  # Lista de meses selecionados
                'empresa': empresa_list,  # Lista de empresas selecionadas
                'filtro_livre': filtro_livre,
            }
        }
        
        return render(request, self.template_name, context)
    
    def processar_dados_hierarquicos_com_datas(self, resultados):
        """Processa os resultados da consulta para criar estrutura hierárquica com datas"""
        hierarquia = {}
        datas_unicas = set()
        
        for i, row in enumerate(resultados):
            data, id_orcamento, cd_empresa, nm_empresa, dre_id, dre_nivel, nivel1, nivel2, nivel3, centro_resultado, valor = row
            
            # Adicionar data ao conjunto de datas únicas
            if data:
                datas_unicas.add(data)
            
            # Organizar por DRE
            if dre_id not in hierarquia:
                hierarquia[dre_id] = {
                    'dre_id': dre_id,
                    'dre_nivel': dre_nivel,
                    'niveis1': {},
                    'total': 0,
                    'valores_por_data': {}
                }
            
            # Organizar por nível 1
            if nivel1 not in hierarquia[dre_id]['niveis1']:
                hierarquia[dre_id]['niveis1'][nivel1] = {
                    'nome': nivel1,
                    'niveis2': {},
                    'total': 0,
                    'valores_por_data': {}
                }
            
            # Organizar por nível 2
            if nivel2 not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2']:
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2] = {
                    'nome': nivel2,
                    'niveis3': {},
                    'total': 0,
                    'valores_por_data': {}
                }
            
            # Adicionar nível 3
            if nivel3 not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3']:
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3] = {
                    'nome': nivel3,
                    'centros_resultado': {},
                    'valor': 0,
                    'valores_por_data': {}
                }
            
            # Adicionar Centro Resultado (nível 4)
            if centro_resultado not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado']:
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado] = {
                    'nome': centro_resultado,
                    'valor': 0,
                    'valores_por_data': {},
                    'orcamento_ids': {}
                }
            
            # Processar valor
            valor_float = float(valor) if valor else 0
            
            # Centro Resultado (nível 4) - somar valores (não sobrescrever)
            if data:
                if data not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado]['valores_por_data']:
                    hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado]['valores_por_data'][data] = 0
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado]['valores_por_data'][data] += valor_float
                
                # Armazenar ID do orçamento (usar o último encontrado)
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado]['orcamento_ids'][data] = id_orcamento
            
            # Somar valores nos níveis superiores
            hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['centros_resultado'][centro_resultado]['valor'] += valor_float
            
            # Nível 3 - somar valores por data
            if data:
                if data not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['valores_por_data']:
                    hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['valores_por_data'][data] = 0
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['valores_por_data'][data] += valor_float
            
            # Somar valores nos níveis superiores
            hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['niveis3'][nivel3]['valor'] += valor_float
            
            # Nível 2 - somar valores por data
            if data:
                if data not in hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['valores_por_data']:
                    hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['valores_por_data'][data] = 0
                hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['valores_por_data'][data] += valor_float
            
            # Nível 1 - somar valores por data
            if data:
                if data not in hierarquia[dre_id]['niveis1'][nivel1]['valores_por_data']:
                    hierarquia[dre_id]['niveis1'][nivel1]['valores_por_data'][data] = 0
                hierarquia[dre_id]['niveis1'][nivel1]['valores_por_data'][data] += valor_float
            
            # Nível DRE - somar valores por data
            if data:
                if data not in hierarquia[dre_id]['valores_por_data']:
                    hierarquia[dre_id]['valores_por_data'][data] = 0
                hierarquia[dre_id]['valores_por_data'][data] += valor_float
            
            # Somar totais
            hierarquia[dre_id]['niveis1'][nivel1]['total'] += valor_float
            hierarquia[dre_id]['niveis1'][nivel1]['niveis2'][nivel2]['total'] += valor_float
            hierarquia[dre_id]['total'] += valor_float
        
        # Ordenar datas
        datas_unicas = sorted(list(datas_unicas))
        
        return hierarquia, datas_unicas

    def estrutura_completa_dre(self, dres, dados_hierarquicos, datas_unicas):
        """Garante que todas as DREs da tabela gestao_dre apareçam na estrutura hierárquica,
        mesmo que não tenham dados orçados associados."""
        estrutura_completa = {}
        
        # Primeiro, criar estrutura vazia para todas as DREs
        for dre in dres:
            dre_id = dre.id
            dre_nivel = dre.nivel
            
            estrutura_completa[dre_id] = {
                'dre_id': dre_id,
                'dre_nivel': dre_nivel,
                'niveis1': {},
                'total': 0,
                'valores_por_data': {},
                'av_por_data': {}  # Análise Vertical por data
            }
            
            # Inicializar valores por data como 0
            for data in datas_unicas:
                estrutura_completa[dre_id]['valores_por_data'][data] = 0
                estrutura_completa[dre_id]['av_por_data'][data] = 0
        
        # Agora adicionar os dados orçados existentes
        for dre_id, dre_data in dados_hierarquicos.items():
            if dre_id in estrutura_completa:
                # Copiar dados existentes
                estrutura_completa[dre_id]['niveis1'] = dre_data['niveis1']
                estrutura_completa[dre_id]['total'] = dre_data['total']
                estrutura_completa[dre_id]['valores_por_data'] = dre_data['valores_por_data']
                
        
        # Calcular somatório para DRE ID 4 (soma dos IDs 2 e 3)
        if 2 in estrutura_completa and 3 in estrutura_completa and 4 in estrutura_completa:
            dre_2 = estrutura_completa[2]
            dre_3 = estrutura_completa[3]
            dre_4 = estrutura_completa[4]
            
            # Somar valores por data
            for data in datas_unicas:
                valor_2 = dre_2['valores_por_data'].get(data, 0)
                valor_3 = dre_3['valores_por_data'].get(data, 0)
                dre_4['valores_por_data'][data] = valor_2 + valor_3
            
            # Somar totais
            dre_4['total'] = dre_2['total'] + dre_3['total']
            
            # Combinar níveis 1 dos IDs 2 e 3
            dre_4['niveis1'] = {}
            
            # Adicionar níveis 1 do ID 2
            for nivel1_nome, nivel1_data in dre_2['niveis1'].items():
                dre_4['niveis1'][nivel1_nome] = nivel1_data
            
            # Adicionar níveis 1 do ID 3
            for nivel1_nome, nivel1_data in dre_3['niveis1'].items():
                if nivel1_nome in dre_4['niveis1']:
                    # Se já existe, somar os valores
                    for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                        if nivel2_nome in dre_4['niveis1'][nivel1_nome]['niveis2']:
                            # Somar níveis 3
                            for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                                if nivel3_nome in dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3']:
                                    dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valor'] += nivel3_data['valor']
                                    # Somar valores por data
                                    for data in datas_unicas:
                                        valor_original = dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'].get(data, 0)
                                        valor_novo = nivel3_data['valores_por_data'].get(data, 0)
                                        dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'][data] = valor_original + valor_novo
                                    
                                    # Somar Centros Resultado (nível 4)
                                    for centro_resultado_nome, centro_resultado_data in nivel3_data['centros_resultado'].items():
                                        if centro_resultado_nome in dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['centros_resultado']:
                                            dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['centros_resultado'][centro_resultado_nome]['valor'] += centro_resultado_data['valor']
                                            # Somar valores por data
                                            for data in datas_unicas:
                                                valor_original = dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['centros_resultado'][centro_resultado_nome]['valores_por_data'].get(data, 0)
                                                valor_novo = centro_resultado_data['valores_por_data'].get(data, 0)
                                                dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['centros_resultado'][centro_resultado_nome]['valores_por_data'][data] = valor_original + valor_novo
                                        else:
                                            dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['centros_resultado'][centro_resultado_nome] = centro_resultado_data
                                else:
                                    dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome] = nivel3_data
                        else:
                            dre_4['niveis1'][nivel1_nome]['niveis2'][nivel2_nome] = nivel2_data
                else:
                    dre_4['niveis1'][nivel1_nome] = nivel1_data
        
        # Calcular somatório para DRE ID 7 (soma dos IDs 4, 5 e 6)
        if 4 in estrutura_completa and 5 in estrutura_completa and 6 in estrutura_completa and 7 in estrutura_completa:
            dre_4 = estrutura_completa[4]
            dre_5 = estrutura_completa[5]
            dre_6 = estrutura_completa[6]
            dre_7 = estrutura_completa[7]
            
            # Somar valores por data
            for data in datas_unicas:
                valor_4 = dre_4['valores_por_data'].get(data, 0)
                valor_5 = dre_5['valores_por_data'].get(data, 0)
                valor_6 = dre_6['valores_por_data'].get(data, 0)
                dre_7['valores_por_data'][data] = valor_4 + valor_5 + valor_6
            
            # Somar totais
            dre_7['total'] = dre_4['total'] + dre_5['total'] + dre_6['total']
            
            # Combinar níveis 1 dos IDs 4, 5 e 6
            dre_7['niveis1'] = {}
            
            # Adicionar níveis 1 do ID 4
            for nivel1_nome, nivel1_data in dre_4['niveis1'].items():
                dre_7['niveis1'][nivel1_nome] = nivel1_data
            
            # Adicionar níveis 1 do ID 5
            for nivel1_nome, nivel1_data in dre_5['niveis1'].items():
                if nivel1_nome in dre_7['niveis1']:
                    # Se já existe, somar os valores
                    for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                        if nivel2_nome in dre_7['niveis1'][nivel1_nome]['niveis2']:
                            # Somar níveis 3
                            for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                                if nivel3_nome in dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3']:
                                    dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valor'] += nivel3_data['valor']
                                    # Somar valores por data
                                    for data in datas_unicas:
                                        valor_original = dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'].get(data, 0)
                                        valor_novo = nivel3_data['valores_por_data'].get(data, 0)
                                        dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'][data] = valor_original + valor_novo
                                else:
                                    dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome] = nivel3_data
                        else:
                            dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome] = nivel2_data
                else:
                    dre_7['niveis1'][nivel1_nome] = nivel1_data
            
            # Adicionar níveis 1 do ID 6
            for nivel1_nome, nivel1_data in dre_6['niveis1'].items():
                if nivel1_nome in dre_7['niveis1']:
                    # Se já existe, somar os valores
                    for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                        if nivel2_nome in dre_7['niveis1'][nivel1_nome]['niveis2']:
                            # Somar níveis 3
                            for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                                if nivel3_nome in dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3']:
                                    dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valor'] += nivel3_data['valor']
                                    # Somar valores por data
                                    for data in datas_unicas:
                                        valor_original = dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'].get(data, 0)
                                        valor_novo = nivel3_data['valores_por_data'].get(data, 0)
                                        dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'][data] = valor_original + valor_novo
                                else:
                                    dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome] = nivel3_data
                        else:
                            dre_7['niveis1'][nivel1_nome]['niveis2'][nivel2_nome] = nivel2_data
                else:
                    dre_7['niveis1'][nivel1_nome] = nivel1_data
        
        # Calcular somatório para DRE ID 9 (soma dos IDs 7 e 8)
        if 7 in estrutura_completa and 8 in estrutura_completa and 9 in estrutura_completa:
            dre_7 = estrutura_completa[7]
            dre_8 = estrutura_completa[8]
            dre_9 = estrutura_completa[9]
            
            # Somar valores por data
            for data in datas_unicas:
                valor_7 = dre_7['valores_por_data'].get(data, 0)
                valor_8 = dre_8['valores_por_data'].get(data, 0)
                dre_9['valores_por_data'][data] = valor_7 + valor_8
            
            # Somar totais
            dre_9['total'] = dre_7['total'] + dre_8['total']
            
            # Combinar níveis 1 dos IDs 7 e 8
            dre_9['niveis1'] = {}
            
            # Adicionar níveis 1 do ID 7
            for nivel1_nome, nivel1_data in dre_7['niveis1'].items():
                dre_9['niveis1'][nivel1_nome] = nivel1_data
            
            # Adicionar níveis 1 do ID 8
            for nivel1_nome, nivel1_data in dre_8['niveis1'].items():
                if nivel1_nome in dre_9['niveis1']:
                    # Se já existe, somar os valores
                    for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                        if nivel2_nome in dre_9['niveis1'][nivel1_nome]['niveis2']:
                            # Somar níveis 3
                            for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                                if nivel3_nome in dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3']:
                                    dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valor'] += nivel3_data['valor']
                                    # Somar valores por data
                                    for data in datas_unicas:
                                        valor_original = dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'].get(data, 0)
                                        valor_novo = nivel3_data['valores_por_data'].get(data, 0)
                                        dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'][data] = valor_original + valor_novo
                                else:
                                    dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome] = nivel3_data
                        else:
                            dre_9['niveis1'][nivel1_nome]['niveis2'][nivel2_nome] = nivel2_data
                else:
                    dre_9['niveis1'][nivel1_nome] = nivel1_data
        
        # Calcular somatório para DRE ID 11 (soma dos IDs 9 e 10)
        if 9 in estrutura_completa and 10 in estrutura_completa and 11 in estrutura_completa:
            dre_9 = estrutura_completa[9]
            dre_10 = estrutura_completa[10]
            dre_11 = estrutura_completa[11]
            
            # Somar valores por data
            for data in datas_unicas:
                valor_9 = dre_9['valores_por_data'].get(data, 0)
                valor_10 = dre_10['valores_por_data'].get(data, 0)
                dre_11['valores_por_data'][data] = valor_9 + valor_10
            
            # Somar totais
            dre_11['total'] = dre_9['total'] + dre_10['total']
            
            # Combinar níveis 1 dos IDs 9 e 10
            dre_11['niveis1'] = {}
            
            # Adicionar níveis 1 do ID 9
            for nivel1_nome, nivel1_data in dre_9['niveis1'].items():
                dre_11['niveis1'][nivel1_nome] = nivel1_data
            
            # Adicionar níveis 1 do ID 10
            for nivel1_nome, nivel1_data in dre_10['niveis1'].items():
                if nivel1_nome in dre_11['niveis1']:
                    # Se já existe, somar os valores
                    for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                        if nivel2_nome in dre_11['niveis1'][nivel1_nome]['niveis2']:
                            # Somar níveis 3
                            for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                                if nivel3_nome in dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3']:
                                    dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valor'] += nivel3_data['valor']
                                    # Somar valores por data
                                    for data in datas_unicas:
                                        valor_original = dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'].get(data, 0)
                                        valor_novo = nivel3_data['valores_por_data'].get(data, 0)
                                        dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome]['valores_por_data'][data] = valor_original + valor_novo
                                else:
                                    dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome]['niveis3'][nivel3_nome] = nivel3_data
                        else:
                            dre_11['niveis1'][nivel1_nome]['niveis2'][nivel2_nome] = nivel2_data
                else:
                    dre_11['niveis1'][nivel1_nome] = nivel1_data
        
        # Calcular Análise Vertical (AV) - percentual em relação ao total da receita
        # A receita total é o somatório do ID 2 da estrutura DRE
        receita_dre_id = 2  # ID 2 é a receita total
        receita_total = 0
        receita_por_data = {}
        
        # Obter dados da receita (DRE ID 2)
        if receita_dre_id in estrutura_completa:
            receita_total = estrutura_completa[receita_dre_id]['total']
            receita_por_data = estrutura_completa[receita_dre_id]['valores_por_data']
        
        # Verificar se há dados de receita em dados_hierarquicos
        for dre_id, dre_data in estrutura_completa.items():
            # Calcular AV para valores por data
            for data in datas_unicas:
                valor = dre_data['valores_por_data'].get(data, 0)
                receita_data = receita_por_data.get(data, 0)
                
                if receita_data != 0:
                    av_percentual = (valor / receita_data) * 100
                else:
                    av_percentual = 0
                dre_data['av_por_data'][data] = av_percentual
            
            # Calcular AV para total
            if receita_total != 0:
                dre_data['av_total'] = (dre_data['total'] / receita_total) * 100
            else:
                dre_data['av_total'] = 0
            
            # Calcular AV para níveis hierárquicos
            for nivel1_nome, nivel1_data in dre_data['niveis1'].items():
                # Inicializar av_por_data se não existir
                if 'av_por_data' not in nivel1_data:
                    nivel1_data['av_por_data'] = {}
                
                # Calcular AV para valores por data no nível 1
                for data in datas_unicas:
                    valor = nivel1_data['valores_por_data'].get(data, 0)
                    receita_data = receita_por_data.get(data, 0)
                    
                    if receita_data != 0:
                        av_percentual = (valor / receita_data) * 100
                    else:
                        av_percentual = 0
                    nivel1_data['av_por_data'][data] = av_percentual
                
                # Calcular AV para total do nível 1
                if receita_total != 0:
                    nivel1_data['av_total'] = (nivel1_data['total'] / receita_total) * 100
                else:
                    nivel1_data['av_total'] = 0
                
                for nivel2_nome, nivel2_data in nivel1_data['niveis2'].items():
                    # Inicializar av_por_data se não existir
                    if 'av_por_data' not in nivel2_data:
                        nivel2_data['av_por_data'] = {}
                    
                    # Calcular AV para valores por data no nível 2
                    for data in datas_unicas:
                        valor = nivel2_data['valores_por_data'].get(data, 0)
                        receita_data = receita_por_data.get(data, 0)
                        
                        if receita_data != 0:
                            av_percentual = (valor / receita_data) * 100
                        else:
                            av_percentual = 0
                        nivel2_data['av_por_data'][data] = av_percentual
                    
                    # Calcular AV para total do nível 2
                    if receita_total != 0:
                        nivel2_data['av_total'] = (nivel2_data['total'] / receita_total) * 100
                    else:
                        nivel2_data['av_total'] = 0
                    
                    for nivel3_nome, nivel3_data in nivel2_data['niveis3'].items():
                        # Inicializar av_por_data se não existir
                        if 'av_por_data' not in nivel3_data:
                            nivel3_data['av_por_data'] = {}
                        
                        # Calcular AV para valores por data no nível 3
                        for data in datas_unicas:
                            valor = nivel3_data['valores_por_data'].get(data, 0)
                            receita_data = receita_por_data.get(data, 0)
                            
                            if receita_data != 0:
                                av_percentual = (valor / receita_data) * 100
                            else:
                                av_percentual = 0
                            nivel3_data['av_por_data'][data] = av_percentual
                        
                        # Calcular AV para valor do nível 3
                        if receita_total != 0:
                            nivel3_data['av_total'] = (nivel3_data['valor'] / receita_total) * 100
                        else:
                            nivel3_data['av_total'] = 0
                        
                        # Calcular AV para Centros Resultado (nível 4)
                        for centro_resultado_nome, centro_resultado_data in nivel3_data['centros_resultado'].items():
                            # Inicializar av_por_data se não existir
                            if 'av_por_data' not in centro_resultado_data:
                                centro_resultado_data['av_por_data'] = {}
                            
                            # Calcular AV para valores por data no Centro Resultado
                            for data in datas_unicas:
                                valor = centro_resultado_data['valores_por_data'].get(data, 0)
                                receita_data = receita_por_data.get(data, 0)
                                
                                if receita_data != 0:
                                    av_percentual = (valor / receita_data) * 100
                                else:
                                    av_percentual = 0
                                centro_resultado_data['av_por_data'][data] = av_percentual
                            
                            # Calcular AV para valor do Centro Resultado
                            if receita_total != 0:
                                centro_resultado_data['av_total'] = (centro_resultado_data['valor'] / receita_total) * 100
                            else:
                                centro_resultado_data['av_total'] = 0
        
        return estrutura_completa


class OrcamentoVisualizacaoView(LoginRequiredMixin, View):
    """View para visualização do orçamento com filtros"""
    template_name = 'Gestao/orcamento_visualizacao.html'
    
    def get(self, request, *args, **kwargs):
        # Verificar permissão para acessar visualização de orçamento
        acesso = AcessoAcoes(request, 'Gestao Orçamento', 'Orçamento')
        if not acesso:
            return render(request, 'Forbidden.html')
        
        # Parâmetros de filtro - suportando múltipla escolha
        empresa_list = request.GET.getlist('empresa')  # Múltipla escolha para empresas
        ano = request.GET.get('ano', datetime.now().year)
        mes_list = request.GET.getlist('mes')  # Múltipla escolha para meses
        conta_id = request.GET.get('conta')
        
        # Buscar dados para os filtros
        # Filtrar apenas empresas que tenham movimentações no orçamento
        empresas = Empresa.objects.filter(
            id__in=Orcamento.objects.values_list('empresa_id', flat=True).distinct()
        ).order_by('empresa')
        contas = Conta.objects.filter(nivel=1).order_by('codigo')  # Apenas contas de nível 1
        
        # Filtrar orçamentos
        orcamentos = Orcamento.objects.all()
        if empresa_list:
            orcamentos = orcamentos.filter(empresa_id__in=empresa_list)
        if ano:
            orcamentos = orcamentos.filter(data__year=ano)
        if mes_list:
            orcamentos = orcamentos.filter(data__month__in=mes_list)
        if conta_id:
            # Filtrar por conta e suas subcontas
            conta = Conta.objects.get(id=conta_id)
            orcamentos = orcamentos.filter(conta__codigo__startswith=conta.codigo)
        
        # Buscar todas as contas para a tabela (hierarquia completa)
        todas_contas = Conta.objects.all().order_by('ordem')
        
        # Preparar dados para a tabela
        meses = []
        dados_tabela = []
        
        # Definir período (ano inteiro ou meses específicos)
        if mes_list:
            # Apenas meses selecionados
            meses = []
            for mes_num in mes_list:
                try:
                    data_mes = datetime(int(ano), int(mes_num), 1).date()
                    meses.append(data_mes)
                except (ValueError, TypeError):
                    continue
            meses.sort()  # Ordenar por data
        else:
            # Ano inteiro
            data_inicio = datetime(int(ano), 1, 1).date()
            data_fim = datetime(int(ano), 12, 31).date()
            meses = [datetime(int(ano), m, 1).date() for m in range(1, 13)]
        
        # Preparar dados para cada conta
        for conta in todas_contas:
            linha = {
                'conta': conta,
                'nivel': conta.nivel,
                'valores_mes': {}
            }
            
            # Buscar valores para cada mês
            for data_mes in meses:
                valor = orcamentos.filter(
                    conta=conta,
                    data__year=data_mes.year,
                    data__month=data_mes.month
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                linha['valores_mes'][data_mes.month] = valor
            
            dados_tabela.append(linha)
        
        # Calcular totais
        totais_mes = {}
        for data_mes in meses:
            total = orcamentos.filter(
                data__year=data_mes.year,
                data__month=data_mes.month
            ).aggregate(total=Sum('valor'))['total'] or 0
            totais_mes[data_mes.month] = total
        
        context = {
            'empresas': empresas,
            'contas': contas,
            'dados_tabela': dados_tabela,
            'meses': meses,
            'totais_mes': totais_mes,
            'filtros': {
                'empresa': empresa_list,  # Lista de empresas selecionadas
                'ano': ano,
                'mes': mes_list,  # Lista de meses selecionados
                'conta_id': conta_id or '',
            },
            'ano_atual': datetime.now().year,
            'mes_atual': datetime.now().month,
        }
        
        return render(request, self.template_name, context)


class ExportOrcamentoExcelView(LoginRequiredMixin, View):
    """View para exportar dados do orçamento para Excel"""
    
    def get(self, request, *args, **kwargs):
        # Verificar permissão para acessar exportação de orçamento
        acesso = AcessoAcoes(request, 'Gestao Orçamento', 'Orçamento')
        if not acesso:
            return JsonResponse({
                'status': 'forbidden', 
                'message': 'Usuário sem permissão para exportar orçamentos'
            })
        
        try:
            # Parâmetros de filtro - suportando múltipla escolha
            empresa_list = request.GET.getlist('empresa')
            ano = request.GET.get('ano', datetime.now().year)
            mes_list = request.GET.getlist('mes')
            conta_id = request.GET.get('conta')
            
            # Filtrar orçamentos
            orcamentos = Orcamento.objects.select_related('empresa', 'conta', 'centro_custo')
            if empresa_list:
                orcamentos = orcamentos.filter(empresa_id__in=empresa_list)
            if ano:
                orcamentos = orcamentos.filter(data__year=ano)
            if mes_list:
                orcamentos = orcamentos.filter(data__month__in=mes_list)
            if conta_id:
                # Filtrar por conta e suas subcontas
                try:
                    conta = Conta.objects.get(id=conta_id)
                    orcamentos = orcamentos.filter(conta__codigo__startswith=conta.codigo)
                except Conta.DoesNotExist:
                    pass
            
            # Buscar todas as contas para a tabela (hierarquia completa)
            todas_contas = Conta.objects.all().order_by('ordem')
            
            # Preparar dados para a tabela
            meses = []
            
            # Definir período (ano inteiro ou meses específicos)
            if mes_list:
                # Apenas meses selecionados
                meses = []
                for mes_num in mes_list:
                    try:
                        data_mes = datetime(int(ano), int(mes_num), 1).date()
                        meses.append(data_mes)
                    except (ValueError, TypeError):
                        continue
                meses.sort()
            else:
                # Ano inteiro
                meses = [datetime(int(ano), m, 1).date() for m in range(1, 13)]
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Orçamento {ano}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            subtitle_font = Font(bold=True, color="FFFFFF", size=10)
            subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"RELATÓRIO DE ORÇAMENTO - ANO {ano}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Informações dos filtros
            row = 2
            ws.cell(row=row, column=1, value="Filtros Aplicados:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
            ws.cell(row=row, column=1, value="Ano:")
            ws.cell(row=row, column=2, value=ano)
            
            row += 1
            if empresa_list:
                empresas_filtradas = Empresa.objects.filter(id__in=empresa_list).values_list('empresa', flat=True)
                ws.cell(row=row, column=1, value="Empresas:")
                ws.cell(row=row, column=2, value=", ".join(empresas_filtradas))
            
            row += 1
            if mes_list:
                nomes_meses = []
                for mes_num in mes_list:
                    try:
                        data_mes = datetime(int(ano), int(mes_num), 1).date()
                        nomes_meses.append(data_mes.strftime('%B'))
                    except:
                        continue
                ws.cell(row=row, column=1, value="Meses:")
                ws.cell(row=row, column=2, value=", ".join(nomes_meses))
            
            row += 2  # Espaço antes da tabela
            
            # Cabeçalhos da tabela
            headers = ['Conta', 'Nível', 'Código']
            for data_mes in meses:
                headers.append(data_mes.strftime('%B/%Y'))
            headers.append('Total Anual')
            
            # Aplicar estilos ao cabeçalho
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            row += 1
            
            # Dados da tabela
            for conta in todas_contas:
                linha = []
                linha.append(conta.nome)
                linha.append(conta.nivel or 0)
                linha.append(conta.codigo)
                
                # Buscar valores para cada mês
                total_anual = 0
                for data_mes in meses:
                    valor = orcamentos.filter(
                        conta=conta,
                        data__year=data_mes.year,
                        data__month=data_mes.month
                    ).aggregate(total=Sum('valor'))['total'] or 0
                    
                    linha.append(valor)
                    total_anual += valor
                
                linha.append(total_anual)
                
                # Adicionar linha ao Excel
                for col, valor in enumerate(linha, 1):
                    cell = ws.cell(row=row, column=col, value=valor)
                    cell.border = border
                    
                    # Formatação para valores monetários
                    if col > 3:  # Colunas de valores
                        if isinstance(valor, (int, float)) and valor != 0:
                            cell.number_format = '#,##0.00'
                            if valor < 0:
                                cell.font = Font(color="FF0000")
                
                row += 1
            
            # Linha de totais
            ws.cell(row=row, column=1, value="TOTAIS")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value="")
            ws.cell(row=row, column=3, value="")
            
            total_geral = 0
            for col, data_mes in enumerate(meses, 4):
                total = orcamentos.filter(
                    data__year=data_mes.year,
                    data__month=data_mes.month
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                cell = ws.cell(row=row, column=col, value=total)
                cell.font = Font(bold=True)
                cell.fill = subtitle_fill
                cell.font = subtitle_font
                cell.border = border
                cell.number_format = '#,##0.00'
                
                total_geral += total
            
            # Total geral
            cell = ws.cell(row=row, column=len(headers), value=total_geral)
            cell.font = Font(bold=True)
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = border
            cell.number_format = '#,##0.00'
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Criar resposta
            filename = f"orcamento_{ano}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error', 
                'message': f'Erro ao gerar arquivo Excel: {str(e)}'
            }, status=500)
